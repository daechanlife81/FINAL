"""
SEM Robustness Check (MLR 재검증) Excel 생성
박사학위논문 마지막 점검 (2026-04-26)

목적: WLSMV 검토 결과 + MLR 본 분석 결과를 별도 표로 보존
파일명: SEM_robustness_check.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

# 공통 스타일
title_font = Font(name="Times New Roman", bold=True, size=13)
header_font = Font(name="Times New Roman", bold=True, size=11)
body_font = Font(name="Times New Roman", size=11)
bold_font = Font(name="Times New Roman", bold=True, size=11)
note_font = Font(name="Times New Roman", italic=True, size=10)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin")
thick = Side(style="medium")
border_top = Border(top=thick, bottom=thin)
border_mid = Border(top=thin, bottom=thin)
border_bottom = Border(bottom=thick)
header_fill = PatternFill("solid", fgColor="E8E8E8")


def add_title(ws, text, ncol):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = title_font
    cell.alignment = center
    ws.row_dimensions[1].height = 22


def write_header(ws, headers, row=3):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.border = border_top
        cell.fill = header_fill


def write_data(ws, rows, start_row=4, num_cols=None, bold_first=False):
    last = start_row
    for r, row_data in enumerate(rows, start_row):
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = bold_font if (bold_first and c == 1) else body_font
            cell.alignment = center if c > 1 else left
            cell.border = border_mid
        last = r
    if num_cols is None:
        num_cols = len(rows[0])
    for c in range(1, num_cols + 1):
        ws.cell(row=last, column=c).border = border_bottom
    return last


def add_note(ws, note, start_row, ncol):
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=ncol)
    cell = ws.cell(row=start_row, column=1, value=note)
    cell.font = note_font
    cell.alignment = left


# ============================================================
# Sheet 1: 추정법 비교 (Estimator Comparison)
# ============================================================
ws1 = wb.create_sheet("1. Estimator Comparison")
add_title(ws1, "추정법 비교: WLSMV vs ML vs MLR (Robustness Check)", 7)

write_header(ws1,
             ["Estimator", "χ²", "df", "CFI", "TLI", "RMSEA", "SRMR"], row=3)

data = [
    ["WLSMV", 3702.17, 447, 0.768, 0.869, 0.074, 0.120],
    ["ML (기존)", 2165.32, 447, 0.881, 0.865, 0.054, 0.074],
    ["MLR (최종 채택)", 2092.84, 447, 0.882, 0.866, 0.054, 0.074],
]
last = write_data(ws1, data, start_row=4, num_cols=7, bold_first=True)

add_note(ws1,
         "주. WLSMV: 분산 음수(-2.295) Heywood case 발생, 폴리코릭 상관 1.0 근접 → 채택 불가",
         last + 2, 7)
add_note(ws1,
         "    MLR: Satorra-Bentler 강건 표준오차, ML과 결과 거의 동일하여 추정법 견고성 확인",
         last + 3, 7)
add_note(ws1,
         "    적합도 기준: Hu & Bentler(1999) 2-지수 전략 (RMSEA ≤ .06, SRMR ≤ .08)",
         last + 4, 7)

ws1.column_dimensions["A"].width = 18
for col in "BCDEFG":
    ws1.column_dimensions[col].width = 12


# ============================================================
# Sheet 2: MLR 적합도 (Model Fit - MLR)
# ============================================================
ws2 = wb.create_sheet("2. MLR Model Fit")
add_title(ws2, "구조모형 적합도 (MLR Robust)", 4)

write_header(ws2, ["지수", "값", "기준", "판단"], row=3)

fit_data = [
    ["χ² (Satorra-Bentler)", "2092.837", "-", "-"],
    ["df", "447", "-", "-"],
    ["p-value", "<.001", "-", "-"],
    ["χ²/df", "4.682", "≤5 수용", "수용"],
    ["CFI (robust)", ".882", "close to .95", "변수 수 고려 시 수용"],
    ["TLI (robust)", ".866", "close to .95", "변수 수 고려 시 수용"],
    ["RMSEA (robust) [90% CI]", ".054 [.051, .056]", "≤.06", "충족"],
    ["SRMR", ".074", "≤.08", "충족"],
]
last = write_data(ws2, fit_data, start_row=4, num_cols=4, bold_first=True)

add_note(ws2,
         "주. MLR(Maximum Likelihood with Robust standard errors) 추정 - Satorra & Bentler(1994)",
         last + 2, 4)
add_note(ws2,
         "    Hu & Bentler(1999)의 2-지수 전략(RMSEA + SRMR) 충족",
         last + 3, 4)
add_note(ws2,
         "    Kenny & McCoach(2003): 변수 수 증가 시 CFI/TLI 악화 경향 → 수용 가능",
         last + 4, 4)

ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 22


# ============================================================
# Sheet 3: MLR 경로계수 (Standardized Path Coefficients)
# ============================================================
ws3 = wb.create_sheet("3. MLR Path Coefficients")
add_title(ws3, "구조경로 계수 (MLR 강건 표준오차)", 6)

write_header(ws3, ["경로", "β", "SE", "p", "유의성", "비교(ML)"], row=3)

# 연구문제 1: 효능감
rq1_header_row = 4
ws3.cell(row=rq1_header_row, column=1, value="[연구문제 1] 독립변수 → 자원봉사 효능감").font = bold_font
ws3.merge_cells(start_row=rq1_header_row, start_column=1,
                end_row=rq1_header_row, end_column=6)

rq1 = [
    ["제도인식 → 효능감", 0.332, 0.029, "<.001", "***", "동일"],
    ["봉사교육경험 → 효능감", 0.156, 0.026, "<.001", "***", "동일"],
    ["주관적 건강 → 효능감", 0.144, 0.030, "<.001", "***", "동일"],
    ["종교활동 → 효능감", 0.079, 0.028, ".005", "**", "동일"],
    ["네트워크 → 효능감", 0.078, 0.025, ".002", "**", "동일"],
    ["교육수준 → 효능감", 0.069, 0.033, ".035", "*", "동일"],
    ["일반적 신뢰 → 효능감", 0.042, 0.030, ".167", "n.s.", "동일"],
    ["사회단체활동 → 효능감", 0.003, 0.027, ".922", "n.s.", "동일"],
]
last = write_data(ws3, rq1, start_row=rq1_header_row + 1, num_cols=6, bold_first=True)

# 연구문제 2: 안녕감
rq2_header_row = last + 2
ws3.cell(row=rq2_header_row, column=1, value="[연구문제 2] 독립변수 → 주관적 안녕감").font = bold_font
ws3.merge_cells(start_row=rq2_header_row, start_column=1,
                end_row=rq2_header_row, end_column=6)

rq2 = [
    ["효능감 → 안녕감", 0.294, 0.035, "<.001", "***", "동일"],
    ["일반적 신뢰 → 안녕감", 0.269, 0.034, "<.001", "***", "동일"],
    ["주관적 건강 → 안녕감", 0.230, 0.036, "<.001", "***", "동일"],
    ["네트워크 → 안녕감", 0.075, 0.028, ".008", "**", "동일"],
    ["사회단체활동 → 안녕감", 0.037, 0.031, ".236", "n.s.", "동일"],
    ["교육수준 → 안녕감", 0.036, 0.039, ".351", "n.s.", "동일"],
    ["종교활동 → 안녕감", -0.022, 0.030, ".456", "n.s.", "동일"],
]
last = write_data(ws3, rq2, start_row=rq2_header_row + 1, num_cols=6, bold_first=True)

# 연구문제 3: 참여
rq3_header_row = last + 2
ws3.cell(row=rq3_header_row, column=1, value="[연구문제 3] 독립변수 → 자원봉사 참여").font = bold_font
ws3.merge_cells(start_row=rq3_header_row, start_column=1,
                end_row=rq3_header_row, end_column=6)

rq3 = [
    ["제도인식 → 참여", 0.231, 0.034, "<.001", "***", "동일"],
    ["사회단체활동 → 참여", 0.097, 0.031, ".002", "**", "동일"],
    ["종교활동 → 참여", 0.091, 0.031, ".003", "**", "동일"],
    ["안녕감 → 참여", 0.083, 0.035, ".018", "*", "동일"],
    ["주관적 건강 → 참여", -0.062, 0.030, ".037", "*", "동일"],
    ["봉사교육경험 → 참여", 0.055, 0.032, ".091", "n.s.", "동일"],
    ["일반적 신뢰 → 참여", 0.032, 0.025, ".207", "n.s.", "동일"],
    ["교육수준 → 참여", -0.028, 0.032, ".383", "n.s.", "동일"],
    ["네트워크 → 참여", 0.000, 0.023, ".996", "n.s.", "동일"],
    ["효능감 → 참여", 0.005, 0.032, ".870", "n.s.", "동일"],
]
last = write_data(ws3, rq3, start_row=rq3_header_row + 1, num_cols=6, bold_first=True)

add_note(ws3, "주. MLR 강건 표준오차 적용 결과", last + 2, 6)
add_note(ws3, "    '비교(ML)' 열: ML 추정 결과와의 일치 여부 - 모든 경로에서 동일하여 추정법 견고성 확인",
         last + 3, 6)
add_note(ws3, "    *** p<.001, ** p<.01, * p<.05, n.s. = not significant",
         last + 4, 6)

ws3.column_dimensions["A"].width = 30
for col in "BCDEF":
    ws3.column_dimensions[col].width = 12


# ============================================================
# Sheet 4: MLR R-squared
# ============================================================
ws4 = wb.create_sheet("4. MLR R-squared")
add_title(ws4, "내생변수 설명력 (R²)", 4)

write_header(ws4, ["내생변수", "R² (MLR)", "R² (ML)", "비고"], row=3)

r2_data = [
    ["자원봉사 효능감", 0.265, 0.265, "동일"],
    ["주관적 안녕감", 0.276, 0.276, "동일"],
    ["자원봉사 참여", 0.202, 0.202, "동일"],
]
last = write_data(ws4, r2_data, start_row=4, num_cols=4, bold_first=True)

add_note(ws4, "주. MLR과 ML 추정법 간 R² 결과 완전 일치 → 추정법 견고성 확인",
         last + 2, 4)

ws4.column_dimensions["A"].width = 22
for col in "BCD":
    ws4.column_dimensions[col].width = 14


# ============================================================
# Sheet 5: 매개효과 (Bootstrap CI)
# ============================================================
ws5 = wb.create_sheet("5. Mediation Bootstrap")
add_title(ws5, "매개효과 검증 (ML + Bootstrap 5,000회 95% CI)", 6)

write_header(ws5, ["경로", "B", "β", "95% CI Lower", "95% CI Upper", "유의성"], row=3)

# 단순매개: 효능감
m1_header = 4
ws5.cell(row=m1_header, column=1,
         value="[단순매개 1] 독립 → 효능감 → 참여 (모든 경로 비유의 → H13 기각)").font = bold_font
ws5.merge_cells(start_row=m1_header, start_column=1, end_row=m1_header, end_column=6)

med1 = [
    ["주관적 건강", 0.0004, 0.0007, -0.0045, 0.0054, "비유의"],
    ["교육수준", 0.0002, 0.0004, -0.0024, 0.0029, "비유의"],
    ["일반적 신뢰", 0.0001, 0.0002, -0.0021, 0.0025, "비유의"],
    ["네트워크", 0.0005, 0.0004, -0.0052, 0.0069, "비유의"],
    ["사회단체활동", 0.0000, 0.0000, -0.0017, 0.0016, "비유의"],
    ["제도인식", 0.0009, 0.0017, -0.0109, 0.0123, "비유의"],
    ["종교활동", 0.0004, 0.0004, -0.0044, 0.0051, "비유의"],
    ["봉사교육경험", 0.0007, 0.0008, -0.0086, 0.0098, "비유의"],
]
last = write_data(ws5, med1, start_row=m1_header + 1, num_cols=6, bold_first=True)

# 단순매개: 안녕감
m2_header = last + 2
ws5.cell(row=m2_header, column=1,
         value="[단순매개 2] 독립 → 안녕감 → 참여 (3개 변수 유의 → H14 부분채택)").font = bold_font
ws5.merge_cells(start_row=m2_header, start_column=1, end_row=m2_header, end_column=6)

med2 = [
    ["주관적 건강", 0.0098, 0.0192, 0.0015, 0.0195, "유의"],
    ["교육수준", 0.0015, 0.0030, -0.0020, 0.0059, "비유의"],
    ["일반적 신뢰", 0.0147, 0.0224, 0.0025, 0.0275, "유의"],
    ["네트워크", 0.0072, 0.0063, 0.0005, 0.0175, "유의"],
    ["사회단체활동", 0.0025, 0.0030, -0.0016, 0.0089, "비유의"],
    ["종교활동", -0.0016, -0.0019, -0.0074, 0.0026, "비유의"],
]
last = write_data(ws5, med2, start_row=m2_header + 1, num_cols=6, bold_first=True)

# 순차매개
m3_header = last + 2
ws5.cell(row=m3_header, column=1,
         value="[순차매개] 독립 → 효능감 → 안녕감 → 참여 (5개 변수 유의)").font = bold_font
ws5.merge_cells(start_row=m3_header, start_column=1, end_row=m3_header, end_column=6)

med3 = [
    ["주관적 건강", 0.0018, 0.0035, 0.0003, 0.0038, "유의"],
    ["교육수준", 0.0009, 0.0017, -0.00001, 0.0023, "비유의"],
    ["일반적 신뢰", 0.0007, 0.0010, -0.0003, 0.0021, "비유의"],
    ["네트워크", 0.0022, 0.0019, 0.0002, 0.0051, "유의"],
    ["사회단체활동", 0.0001, 0.0001, -0.0012, 0.0014, "비유의"],
    ["제도인식", 0.0045, 0.0081, 0.0007, 0.0089, "유의"],
    ["종교활동", 0.0017, 0.0019, 0.0001, 0.0041, "유의"],
    ["봉사교육경험", 0.0034, 0.0038, 0.0005, 0.0070, "유의"],
]
last = write_data(ws5, med3, start_row=m3_header + 1, num_cols=6, bold_first=True)

add_note(ws5, "주. ML 추정 + Bootstrap 5,000회 백분위 95% 신뢰구간",
         last + 2, 6)
add_note(ws5, "    MLR 추정은 lavaan에서 부트스트랩과 동시 적용 불가하여 ML로 별도 검증",
         last + 3, 6)
add_note(ws5, "    95% CI에 0이 포함되지 않으면 유의 (Shrout & Bolger, 2002)",
         last + 4, 6)

ws5.column_dimensions["A"].width = 16
for col in "BCDEF":
    ws5.column_dimensions[col].width = 14


# ============================================================
# Sheet 6: 가설검증 종합 (MLR 결과 기준)
# ============================================================
ws6 = wb.create_sheet("6. Hypothesis Test Summary")
add_title(ws6, "가설검증 종합 (MLR 재검증 결과)", 4)

write_header(ws6, ["가설", "내용", "MLR 판단", "ML 비교"], row=3)

hyp_data = [
    ["H1", "인간자본(건강, 교육) → 효능감", "채택", "동일"],
    ["H2", "사회자본(신뢰, 네트워크, 사회단체) → 효능감", "부분채택", "동일"],
    ["H3", "문화자본(제도인식, 종교, 봉사교육) → 효능감", "채택", "동일"],
    ["H4", "인간자본(건강, 교육) → 안녕감", "부분채택", "동일"],
    ["H5", "사회자본(신뢰, 네트워크, 사회단체) → 안녕감", "부분채택", "동일"],
    ["H6", "종교활동 → 안녕감", "기각", "동일"],
    ["H7", "효능감 → 안녕감", "채택", "동일"],
    ["H8", "인간자본(건강, 교육) → 참여", "기각", "동일"],
    ["H9", "사회자본(신뢰, 네트워크, 사회단체) → 참여", "부분채택", "동일"],
    ["H10", "문화자본(제도인식, 종교, 봉사교육) → 참여", "채택", "동일"],
    ["H11", "효능감 → 참여", "기각", "동일"],
    ["H12", "안녕감 → 참여", "채택", "동일"],
    ["H13", "효능감: 자본→참여 단순매개", "기각", "동일"],
    ["H14", "안녕감: 인간자본·사회자본·종교활동→참여 단순매개", "부분채택", "동일"],
]
last = write_data(ws6, hyp_data, start_row=4, num_cols=4, bold_first=True)

# 종합
summary_row = last + 1
ws6.merge_cells(start_row=summary_row, start_column=1,
                end_row=summary_row, end_column=4)
cell = ws6.cell(row=summary_row, column=1,
                value="종합: 채택 5개, 부분채택 5개, 기각 4개 (총 14개) — MLR과 ML 결과 완전 일치")
cell.font = bold_font
cell.alignment = center

add_note(ws6, "주. MLR 추정법으로 재검증한 결과, ML 추정과 모든 가설 판정이 동일하여 추정법 견고성 확인",
         summary_row + 2, 4)
add_note(ws6, "    순차매개(독립→효능감→안녕감→참여)는 5개 변수에서 유의 (건강, 네트워크, 제도인식, 종교, 봉사교육)",
         summary_row + 3, 4)

ws6.column_dimensions["A"].width = 8
ws6.column_dimensions["B"].width = 50
ws6.column_dimensions["C"].width = 14
ws6.column_dimensions["D"].width = 12


# ============================================================
# Sheet 7: WLSMV 검토 기록 (Diagnostic Record)
# ============================================================
ws7 = wb.create_sheet("7. WLSMV Diagnostic")
add_title(ws7, "WLSMV 추정 검토 기록 (Diagnostic Record)", 3)

write_header(ws7, ["진단 항목", "결과", "판단"], row=3)

diag_data = [
    ["폴리코릭 상관 추정", "re_A_01_1과 B_05_2 간 상관 1.0 근접", "수치적 불안정"],
    ["분산 추정 (re_A_01_1)", "-2.295 (음수 분산)", "Heywood case 발생"],
    ["CFI (scaled)", ".768", "수용 기준(.90) 미달"],
    ["TLI (scaled)", ".869", "수용 기준(.90) 미달"],
    ["RMSEA (scaled)", ".074", "수용 기준 충족"],
    ["SRMR", ".120", "수용 기준(.08) 초과"],
    ["전체 판정", "채택 불가", "MLR로 대체"],
]
last = write_data(ws7, diag_data, start_row=4, num_cols=3, bold_first=True)

add_note(ws7, "주. WLSMV(Weighted Least Squares Mean and Variance adjusted; Muthén, 1984)",
         last + 2, 3)
add_note(ws7, "    이분형 종속변수(re_A_01_1) 정식 추정법이지만 폴리코릭 상관 1.0 근접 및 음수 분산으로 채택 불가",
         last + 3, 3)
add_note(ws7, "    재현 스크립트: scripts/R/sem_wlsmv_check.R",
         last + 4, 3)

ws7.column_dimensions["A"].width = 28
ws7.column_dimensions["B"].width = 36
ws7.column_dimensions["C"].width = 22


# 저장
wb.save("results/SEM_robustness_check.xlsx")
print("저장 완료: results/SEM_robustness_check.xlsx")
print("총 7개 시트:")
print("  1. Estimator Comparison (3개 추정법 비교)")
print("  2. MLR Model Fit (적합도)")
print("  3. MLR Path Coefficients (경로계수)")
print("  4. MLR R-squared (설명력)")
print("  5. Mediation Bootstrap (매개효과)")
print("  6. Hypothesis Test Summary (가설검증)")
print("  7. WLSMV Diagnostic (검토 기록)")
