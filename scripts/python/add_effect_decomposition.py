from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

wb = load_workbook("SEM_results_v2.xlsx")

header_font = Font(name="Times New Roman", bold=True, size=11)
body_font = Font(name="Times New Roman", size=11)
bold_font = Font(name="Times New Roman", bold=True, size=11)
title_font = Font(name="Times New Roman", bold=True, size=13)
note_font = Font(name="Times New Roman", italic=True, size=10)
center = Alignment(horizontal="center", vertical="center")
center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center")
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin")
thick = Side(style="medium")
border_top = Border(top=thick, bottom=thin)
border_mid = Border(top=thin, bottom=thin)
border_bottom = Border(bottom=thick)

ws = wb.create_sheet("효과분해")

# 제목
ws.merge_cells("A1:H1")
ws["A1"] = "구조모형의 인과경로별 효과분해 (N=1,328)"
ws["A1"].font = title_font
ws["A1"].alignment = center

# 헤더 1행
headers_r3 = [("A3", "구분"), ("B3", "독립변수"), ("C3", "직접효과(\u03b2)"),
              ("D3", "간접효과(\u03b2)"), ("H3", "총효과(\u03b2)")]
# 간접효과 병합
ws.merge_cells("D3:G3")

for cell_ref, val in headers_r3:
    cell = ws[cell_ref]
    cell.value = val
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

# H3도 border
ws["H3"].font = header_font
ws["H3"].alignment = center
ws["H3"].border = border_top
# E3, F3, G3 border
for col in ["E", "F", "G"]:
    ws[f"{col}3"].border = border_top

# 헤더 2행
headers_r4 = [("A4", ""), ("B4", ""), ("C4", ""), ("D4", "(a)"), ("E4", "(b)"), ("F4", "(c)"), ("G4", "\ucd1d\uac04\uc811\ud6a8\uacfc"), ("H4", "")]
for cell_ref, val in headers_r4:
    cell = ws[cell_ref]
    cell.value = val
    cell.font = header_font
    cell.alignment = center
    cell.border = border_mid

# 데이터
data = [
    ("\uc778\uac04\uc790\ubcf8", "\uc8fc\uad00\uc801 \uac74\uac15", "-.062", ".001", ".019*", ".004*", ".023*", "-.038"),
    ("", "\uad50\uc721\uc218\uc900", "-.028", ".000", ".003", ".002", ".005", "-.023"),
    ("\uc0ac\ud68c\uc790\ubcf8", "\uc77c\ubc18\uc801 \uc2e0\ub8b0", ".032", ".000", ".022*", ".001", ".024*", ".056"),
    ("", "\ub124\ud2b8\uc6cc\ud06c", "-.000", ".000", ".006*", ".002*", ".009*", ".008"),
    ("", "\uc0ac\ud68c\ub2e8\uccb4\ud65c\ub3d9", ".097", ".000", ".003", ".000", ".003", ".100"),
    ("\ubb38\ud654\uc790\ubcf8", "\uc81c\ub3c4\uc778\uc2dd", ".231", ".002", "-", ".008*", ".010", ".241"),
    ("", "\uc885\uad50\ud65c\ub3d9", ".091", ".000", "-.002", ".002*", ".000", ".092"),
    ("", "\ubd09\uc0ac\uad50\uc721\uacbd\ud5d8", ".055", ".001", "-", ".004*", ".005", ".059"),
]

r = 5
prev_group = None
group_start = None

for grp, var, direct, a, b, c, total_ind, total in data:
    if grp and grp != prev_group:
        if prev_group and group_start and group_start < r:
            ws.merge_cells(start_row=group_start, start_column=1, end_row=r-1, end_column=1)
            ws.cell(row=group_start, column=1).alignment = center_wrap
        group_start = r
        prev_group = grp

    vals = [grp if grp else "", var, direct, a, b, c, total_ind, total]
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=col, value=v)
        cell.font = body_font
        cell.alignment = center
        cell.border = border_mid

    r += 1

# 마지막 그룹 병합
if group_start and group_start < r:
    ws.merge_cells(start_row=group_start, start_column=1, end_row=r-1, end_column=1)
    ws.cell(row=group_start, column=1).alignment = center_wrap

last_row = r - 1
for c in range(1, 9):
    ws.cell(row=last_row, column=c).border = border_bottom

# 주석
note_row = last_row + 1
ws.merge_cells(f"A{note_row}:H{note_row}")
ws[f"A{note_row}"] = "\uc8fc. Bootstrap 5,000\ud68c, Percentile 95% CI. * p<.05 (95% CI\uc5d0 0 \ubbf8\ud3ec\ud568)."
ws[f"A{note_row}"].font = note_font

note_row2 = note_row + 1
ws.merge_cells(f"A{note_row2}:H{note_row2}")
ws[f"A{note_row2}"] = "    (a) X\u2192\ud6a8\ub2a5\uac10(M1)\u2192\ucc38\uc5ec(Y); (b) X\u2192\uc548\ub155\uac10(M2)\u2192\ucc38\uc5ec(Y); (c) X\u2192\ud6a8\ub2a5\uac10(M1)\u2192\uc548\ub155\uac10(M2)\u2192\ucc38\uc5ec(Y)."
ws[f"A{note_row2}"].font = note_font

note_row3 = note_row2 + 1
ws.merge_cells(f"A{note_row3}:H{note_row3}")
ws[f"A{note_row3}"] = "    - : \uacbd\ub85c \ubbf8\uc124\uc815 (\uc81c\ub3c4\uc778\uc2dd, \ubd09\uc0ac\uad50\uc721\uacbd\ud5d8\uc740 \uc548\ub155\uac10\uc73c\ub85c\uc758 \uc9c1\uc811\uacbd\ub85c \uc5c6\uc74c)."
ws[f"A{note_row3}"].font = note_font

# ========== 간접효과 CI 상세 시트 ==========
ws2 = wb.create_sheet("간접효과 CI 상세")

ws2.merge_cells("A1:F1")
ws2["A1"] = "\uac04\uc811\ud6a8\uacfc 95% \uc2e0\ub8b0\uad6c\uac04 \uc0c1\uc138 (Bootstrap 5,000\ud68c)"
ws2["A1"].font = title_font
ws2["A1"].alignment = center

for i, h in enumerate(["\ub3c5\ub9bd\ubcc0\uc218", "\uacbd\ub85c \uc720\ud615", "\u03b2", "95% CI \ud558\ud55c", "95% CI \uc0c1\ud55c", ""], 1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

ci_data = [
    ("\uc8fc\uad00\uc801\uac74\uac15", "(a) \ud6a8\ub2a5\uac10\u2192\ucc38\uc5ec", ".001", "-.004", ".005", ""),
    ("", "(b) \uc548\ub155\uac10\u2192\ucc38\uc5ec", ".019", ".001", ".019", "*"),
    ("", "(c) \ud6a8\ub2a5\uac10\u2192\uc548\ub155\uac10\u2192\ucc38\uc5ec", ".004", ".000", ".004", "*"),
    ("", "\ucd1d\uac04\uc811\ud6a8\uacfc", ".023", ".002", ".023", "*"),
    ("\uad50\uc721\uc218\uc900", "(a) \ud6a8\ub2a5\uac10\u2192\ucc38\uc5ec", ".000", "-.002", ".003", ""),
    ("", "(b) \uc548\ub155\uac10\u2192\ucc38\uc5ec", ".003", "-.002", ".006", ""),
    ("", "(c) \ud6a8\ub2a5\uac10\u2192\uc548\ub155\uac10\u2192\ucc38\uc5ec", ".002", "-.000", ".002", ""),
    ("", "\ucd1d\uac04\uc811\ud6a8\uacfc", ".005", "-.002", ".008", ""),
    ("\uc77c\ubc18\uc801\uc2e0\ub8b0", "(a) \ud6a8\ub2a5\uac10\u2192\ucc38\uc5ec", ".000", "-.002", ".003", ""),
    ("", "(b) \uc548\ub155\uac10\u2192\ucc38\uc5ec", ".022", ".002", ".027", "*"),
    ("", "(c) \ud6a8\ub2a5\uac10\u2192\uc548\ub155\uac10\u2192\ucc38\uc5ec", ".001", "-.000", ".002", ""),
    ("", "\ucd1d\uac04\uc811\ud6a8\uacfc", ".024", ".003", ".028", "*"),
    ("\ub124\ud2b8\uc6cc\ud06c", "(a) \ud6a8\ub2a5\uac10\u2192\ucc38\uc5ec", ".000", "-.005", ".007", ""),
    ("", "(b) \uc548\ub155\uac10\u2192\ucc38\uc5ec", ".006", ".000", ".017", "*"),
    ("", "(c) \ud6a8\ub2a5\uac10\u2192\uc548\ub155\uac10\u2192\ucc38\uc5ec", ".002", ".000", ".005", "*"),
    ("", "\ucd1d\uac04\uc811\ud6a8\uacfc", ".009", ".001", ".022", "*"),
    ("\uc0ac\ud68c\ub2e8\uccb4\ud65c\ub3d9", "(a) \ud6a8\ub2a5\uac10\u2192\ucc38\uc5ec", ".000", "-.002", ".002", ""),
    ("", "(b) \uc548\ub155\uac10\u2192\ucc38\uc5ec", ".003", "-.002", ".009", ""),
    ("", "(c) \ud6a8\ub2a5\uac10\u2192\uc548\ub155\uac10\u2192\ucc38\uc5ec", ".000", "-.001", ".001", ""),
    ("", "\ucd1d\uac04\uc811\ud6a8\uacfc", ".003", "-.002", ".009", ""),
    ("\uc81c\ub3c4\uc778\uc2dd", "(a) \ud6a8\ub2a5\uac10\u2192\ucc38\uc5ec", ".002", "-.010", ".013", ""),
    ("", "(b) -", "-", "-", "-", ""),
    ("", "(c) \ud6a8\ub2a5\uac10\u2192\uc548\ub155\uac10\u2192\ucc38\uc5ec", ".008", ".001", ".009", "*"),
    ("", "\ucd1d\uac04\uc811\ud6a8\uacfc", ".010", "-.005", ".017", ""),
    ("\uc885\uad50\ud65c\ub3d9", "(a) \ud6a8\ub2a5\uac10\u2192\ucc38\uc5ec", ".000", "-.004", ".005", ""),
    ("", "(b) \uc548\ub155\uac10\u2192\ucc38\uc5ec", "-.002", "-.007", ".003", ""),
    ("", "(c) \ud6a8\ub2a5\uac10\u2192\uc548\ub155\uac10\u2192\ucc38\uc5ec", ".002", ".000", ".004", "*"),
    ("", "\ucd1d\uac04\uc811\ud6a8\uacfc", ".000", "-.007", ".007", ""),
    ("\ubd09\uc0ac\uad50\uc721\uacbd\ud5d8", "(a) \ud6a8\ub2a5\uac10\u2192\ucc38\uc5ec", ".001", "-.008", ".010", ""),
    ("", "(b) -", "-", "-", "-", ""),
    ("", "(c) \ud6a8\ub2a5\uac10\u2192\uc548\ub155\uac10\u2192\ucc38\uc5ec", ".004", ".000", ".007", "*"),
    ("", "\ucd1d\uac04\uc811\ud6a8\uacfc", ".005", "-.004", ".013", ""),
]

r = 4
prev_var = None
var_start = None

for var, path_type, beta, ci_lo, ci_hi, sig in ci_data:
    if var and var != prev_var:
        if prev_var and var_start and var_start < r:
            ws2.merge_cells(start_row=var_start, start_column=1, end_row=r-1, end_column=1)
            ws2.cell(row=var_start, column=1).alignment = center_wrap
        var_start = r
        prev_var = var

    vals = [var if var else "", path_type, beta, ci_lo, ci_hi, sig]
    for col, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=col, value=v)
        cell.font = body_font
        cell.alignment = center
        cell.border = border_mid

    r += 1

if var_start and var_start < r:
    ws2.merge_cells(start_row=var_start, start_column=1, end_row=r-1, end_column=1)
    ws2.cell(row=var_start, column=1).alignment = center_wrap

last_row2 = r - 1
for c in range(1, 7):
    ws2.cell(row=last_row2, column=c).border = border_bottom

note_row_ci = last_row2 + 1
ws2.merge_cells(f"A{note_row_ci}:F{note_row_ci}")
ws2[f"A{note_row_ci}"] = "\uc8fc. * 95% \uc2e0\ub8b0\uad6c\uac04\uc5d0 0\uc774 \ud3ec\ud568\ub418\uc9c0 \uc54a\uc73c\uba74 \uc720\uc758\ud55c \uac04\uc811\ud6a8\uacfc. Bootstrap 5,000\ud68c, Percentile CI."
ws2[f"A{note_row_ci}"].font = note_font

# 열 너비
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 10
ws.column_dimensions["F"].width = 10
ws.column_dimensions["G"].width = 12
ws.column_dimensions["H"].width = 12

ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 26
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 6

wb.save("SEM_results_v2.xlsx")
print("saved: SEM_results_v2.xlsx (효과분해, 간접효과 CI 상세 시트 추가)")
