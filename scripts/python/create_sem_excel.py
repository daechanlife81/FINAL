from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

wb = Workbook()

header_font = Font(name="Times New Roman", bold=True, size=11)
body_font = Font(name="Times New Roman", size=11)
title_font = Font(name="Times New Roman", bold=True, size=13)
note_font = Font(name="Times New Roman", italic=True, size=10)
center = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin")
thick = Side(style="medium")
border_top = Border(top=thick, bottom=thin)
border_mid = Border(top=thin, bottom=thin)
border_bottom = Border(bottom=thick)
border_top_bottom = Border(top=thick, bottom=thick)

# ========== Sheet 1: 구조모형 적합도 ==========
ws1 = wb.active
ws1.title = "구조모형 적합도"

ws1.merge_cells("A1:C1")
ws1["A1"] = "구조모형 적합도 검증 결과"
ws1["A1"].font = title_font
ws1["A1"].alignment = center

for i, h in enumerate(["적합도 지수", "값", "기준"], 1):
    cell = ws1.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

fit_data = [
    ("χ²(df)", "2165.325(447), p<.001", "-"),
    ("χ²/df", "4.844", "≤3 양호, ≤5 수용"),
    ("GFI", ".872", "≥.90"),
    ("CFI", ".881", "≥.90"),
    ("TLI", ".865", "≥.90"),
    ("RMSEA", ".054 [.052, .056]", "≤.08"),
    ("SRMR", ".074", "≤.08"),
]

for r, (a, b, c) in enumerate(fit_data, 4):
    for col, v in enumerate([a, b, c], 1):
        cell = ws1.cell(row=r, column=col, value=v)
        cell.font = body_font
        cell.alignment = center
        cell.border = border_mid

for c in range(1, 4):
    ws1.cell(row=10, column=c).border = border_bottom

ws1.merge_cells("A11:C11")
ws1["A11"] = "주. RMSEA, SRMR은 기준 충족. GFI, CFI, TLI는 기준에 근접하며 대표본(N=1,328) 효과 고려 시 수용 가능."
ws1["A11"].font = note_font
ws1.column_dimensions["A"].width = 16
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 30

# ========== Sheet 2: 구조경로 계수 ==========
ws2 = wb.create_sheet("구조경로 계수")

ws2.merge_cells("A1:F1")
ws2["A1"] = "구조모형 경로계수 추정 결과"
ws2["A1"].font = title_font
ws2["A1"].alignment = center

headers2 = ["구분", "경로", "β", "S.E.", "p", "유의도"]
for i, h in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

path_data = [
    ("연구문제1\n독립→효능감", "주관적 건강 → 효능감", ".144", ".030", "<.001", "***"),
    ("", "교육수준 → 효능감", ".069", ".032", ".034", "*"),
    ("", "일반적 신뢰 → 효능감", ".042", ".031", ".171", ""),
    ("", "네트워크 → 효능감", ".078", ".025", ".002", "**"),
    ("", "사회단체활동 → 효능감", ".003", ".027", ".922", ""),
    ("", "제도인식 → 효능감", ".332", ".029", "<.001", "***"),
    ("", "종교활동 → 효능감", ".079", ".029", ".005", "**"),
    ("", "봉사교육경험 → 효능감", ".156", ".026", "<.001", "***"),
    ("연구문제2\n독립→안녕감", "주관적 건강 → 안녕감", ".230", ".036", "<.001", "***"),
    ("", "교육수준 → 안녕감", ".036", ".040", ".364", ""),
    ("", "일반적 신뢰 → 안녕감", ".269", ".034", "<.001", "***"),
    ("", "네트워크 → 안녕감", ".075", ".029", ".009", "**"),
    ("", "사회단체활동 → 안녕감", ".037", ".031", ".242", ""),
    ("", "종교활동 → 안녕감", "-.022", ".030", ".459", ""),
    ("연구문제3\n독립→참여", "주관적 건강 → 참여", "-.062", ".030", ".041", "*"),
    ("", "교육수준 → 참여", "-.028", ".032", ".380", ""),
    ("", "일반적 신뢰 → 참여", ".032", ".025", ".204", ""),
    ("", "네트워크 → 참여", "-.000", ".023", ".996", ""),
    ("", "사회단체활동 → 참여", ".097", ".031", ".002", "**"),
    ("", "제도인식 → 참여", ".231", ".034", "<.001", "***"),
    ("", "종교활동 → 참여", ".091", ".031", ".003", "**"),
    ("", "봉사교육경험 → 참여", ".055", ".032", ".093", ""),
    ("연구문제4\n매개경로", "효능감 → 안녕감", ".294", ".035", "<.001", "***"),
    ("", "효능감 → 참여", ".005", ".032", ".872", ""),
    ("", "안녕감 → 참여", ".083", ".035", ".019", "*"),
    ("통제변수", "과거봉사경험 → 효능감", ".191", "-", "<.001", "***"),
    ("", "과거봉사경험 → 참여", ".242", "-", "<.001", "***"),
]

for r, (grp, path, beta, se, p, sig) in enumerate(path_data, 4):
    for col, v in enumerate([grp, path, beta, se, p, sig], 1):
        cell = ws2.cell(row=r, column=col, value=v)
        cell.font = body_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_mid

last_row = 4 + len(path_data) - 1
for c in range(1, 7):
    ws2.cell(row=last_row, column=c).border = border_bottom

# 구분 열 병합
ws2.merge_cells("A4:A11")   # 연구문제1
ws2.merge_cells("A12:A17")  # 연구문제2
ws2.merge_cells("A18:A25")  # 연구문제3
ws2.merge_cells("A26:A28")  # 연구문제4
ws2.merge_cells("A29:A30")  # 통제변수
for r in [4, 12, 18, 26, 29]:
    ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# R-squared
r2_row = last_row + 2
ws2.merge_cells(f"A{r2_row}:F{r2_row}")
ws2[f"A{r2_row}"] = "설명력: 효능감 R²=.265, 안녕감 R²=.276, 참여 R²=.202"
ws2[f"A{r2_row}"].font = Font(name="Times New Roman", bold=True, size=11)

note_row = r2_row + 1
ws2.merge_cells(f"A{note_row}:F{note_row}")
ws2[f"A{note_row}"] = "주. *p<.05, **p<.01, ***p<.001. 유의하지 않은 통제변수는 생략."
ws2[f"A{note_row}"].font = note_font

ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 28
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 10
ws2.column_dimensions["E"].width = 10
ws2.column_dimensions["F"].width = 10

# ========== Sheet 3: 매개효과 ==========
ws3 = wb.create_sheet("매개효과")

ws3.merge_cells("A1:F1")
ws3["A1"] = "매개효과 검증 결과 (Bootstrap 5,000회, 95% CI)"
ws3["A1"].font = title_font
ws3["A1"].alignment = center

headers3 = ["매개유형", "경로", "β", "95% CI 하한", "95% CI 상한", "유의"]
for i, h in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

med_data = [
    ("단순매개\n독립→효능감→참여", "주관적건강→효능감→참여", ".001", "-.005", ".005", "비유의"),
    ("", "교육수준→효능감→참여", ".000", "-.002", ".003", "비유의"),
    ("", "일반적신뢰→효능감→참여", ".000", "-.002", ".003", "비유의"),
    ("", "네트워크→효능감→참여", ".000", "-.005", ".007", "비유의"),
    ("", "사회단체활동→효능감→참여", ".000", "-.002", ".002", "비유의"),
    ("", "제도인식→효능감→참여", ".002", "-.011", ".013", "비유의"),
    ("", "종교활동→효능감→참여", ".000", "-.004", ".005", "비유의"),
    ("", "봉사교육경험→효능감→참여", ".001", "-.008", ".010", "비유의"),
    ("단순매개\n독립→안녕감→참여", "주관적건강→안녕감→참여", ".019", ".001", ".020", "유의"),
    ("", "교육수준→안녕감→참여", ".003", "-.002", ".006", "비유의"),
    ("", "일반적신뢰→안녕감→참여", ".022", ".002", ".028", "유의"),
    ("", "네트워크→안녕감→참여", ".006", ".000", ".017", "유의"),
    ("", "사회단체활동→안녕감→참여", ".003", "-.002", ".009", "비유의"),
    ("", "종교활동→안녕감→참여", "-.002", "-.008", ".003", "비유의"),
    ("순차매개\n독립→효능감\n→안녕감→참여", "주관적건강→효능감→안녕감→참여", ".004", ".000", ".004", "유의"),
    ("", "교육수준→효능감→안녕감→참여", ".002", "-.000", ".002", "비유의"),
    ("", "일반적신뢰→효능감→안녕감→참여", ".001", "-.000", ".002", "비유의"),
    ("", "네트워크→효능감→안녕감→참여", ".002", ".000", ".005", "유의"),
    ("", "사회단체활동→효능감→안녕감→참여", ".000", "-.001", ".001", "비유의"),
    ("", "제도인식→효능감→안녕감→참여", ".008", ".001", ".009", "유의"),
    ("", "종교활동→효능감→안녕감→참여", ".002", ".000", ".004", "유의"),
    ("", "봉사교육경험→효능감→안녕감→참여", ".004", ".000", ".007", "유의"),
]

for r, (mtype, path, beta, ci_lo, ci_hi, sig) in enumerate(med_data, 4):
    for col, v in enumerate([mtype, path, beta, ci_lo, ci_hi, sig], 1):
        cell = ws3.cell(row=r, column=col, value=v)
        cell.font = body_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_mid

last_row3 = 4 + len(med_data) - 1
for c in range(1, 7):
    ws3.cell(row=last_row3, column=c).border = border_bottom

# 매개유형 열 병합
ws3.merge_cells("A4:A11")
ws3.merge_cells("A12:A17")
ws3.merge_cells("A18:A25")
for r in [4, 12, 18]:
    ws3.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

note_row3 = last_row3 + 1
ws3.merge_cells(f"A{note_row3}:F{note_row3}")
ws3[f"A{note_row3}"] = "주. 95% 신뢰구간에 0이 포함되지 않으면 유의한 매개효과."
ws3[f"A{note_row3}"].font = note_font

ws3.column_dimensions["A"].width = 18
ws3.column_dimensions["B"].width = 36
ws3.column_dimensions["C"].width = 10
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 10

wb.save("구조모형_분석결과.xlsx")
print("저장 완료: 구조모형_분석결과.xlsx")
