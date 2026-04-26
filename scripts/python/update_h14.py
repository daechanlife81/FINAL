from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

wb = load_workbook("SEM_results_v2.xlsx")

if "가설검증 종합" in wb.sheetnames:
    del wb["가설검증 종합"]

header_font = Font(name="Times New Roman", bold=True, size=11)
body_font = Font(name="Times New Roman", size=11)
bold_font = Font(name="Times New Roman", bold=True, size=11)
title_font = Font(name="Times New Roman", bold=True, size=13)
note_font = Font(name="Times New Roman", italic=True, size=10)
center = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin")
thick = Side(style="medium")
border_top = Border(top=thick, bottom=thin)
border_mid = Border(top=thin, bottom=thin)
border_bottom = Border(bottom=thick)

ws = wb.create_sheet("가설검증 종합")

ws.merge_cells("A1:C1")
ws["A1"] = "연구가설 검증 결과 종합 (14개 가설)"
ws["A1"].font = title_font
ws["A1"].alignment = center

for i, h in enumerate(["가설", "내용", "판단"], 1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

hyp_data = [
    ("H1", "인간자본(건강, 교육) \u2192 효능감", "채택"),
    ("H2", "사회자본(신뢰, 네트워크, 사회단체) \u2192 효능감", "부분채택"),
    ("H3", "문화자본(제도인식, 종교, 봉사교육) \u2192 효능감", "채택"),
    ("H4", "인간자본(건강, 교육) \u2192 안녕감", "부분채택"),
    ("H5", "사회자본(신뢰, 네트워크, 사회단체) \u2192 안녕감", "부분채택"),
    ("H6", "종교활동 \u2192 안녕감", "기각"),
    ("H7", "효능감 \u2192 안녕감", "채택"),
    ("H8", "인간자본(건강, 교육) \u2192 참여", "기각"),
    ("H9", "사회자본(신뢰, 네트워크, 사회단체) \u2192 참여", "부분채택"),
    ("H10", "문화자본(제도인식, 종교, 봉사교육) \u2192 참여", "채택"),
    ("H11", "효능감 \u2192 참여", "기각"),
    ("H12", "안녕감 \u2192 참여", "채택"),
    ("H13", "효능감: 자본\u2192참여 단순매개", "기각"),
    ("H14", "안녕감: 인간자본\xb7사회자본\xb7종교활동\u2192참여 단순매개", "부분채택"),
]

for r, (h, content, result) in enumerate(hyp_data, 4):
    for c, v in enumerate([h, content, result], 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = bold_font if c == 3 else body_font
        cell.alignment = center
        cell.border = border_mid

last_row = 4 + len(hyp_data) - 1
for c in range(1, 4):
    ws.cell(row=last_row, column=c).border = border_bottom

summary_row = last_row + 1
ws.merge_cells(f"A{summary_row}:C{summary_row}")
ws[f"A{summary_row}"] = "종합: 채택 5개, 부분채택 5개, 기각 4개 (총 14개)"
ws[f"A{summary_row}"].font = bold_font
ws[f"A{summary_row}"].alignment = center

note_row = summary_row + 1
ws.merge_cells(f"A{note_row}:C{note_row}")
ws[f"A{note_row}"] = "주. 순차매개(독립\u2192효능감\u2192안녕감\u2192참여)는 가설로 설정하지 않고 효과분해 결과로 보고."
ws[f"A{note_row}"].font = note_font

note_row2 = note_row + 1
ws.merge_cells(f"A{note_row2}:C{note_row2}")
ws[f"A{note_row2}"] = "    효과분해 결과, 순차매개 경로는 건강, 네트워크, 제도인식, 종교, 봉사교육의 5개 변수에서 유의."
ws[f"A{note_row2}"].font = note_font

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 14

wb.save("SEM_results_v2.xlsx")
print("saved: SEM_results_v2.xlsx (H14 수정, 종합 수치 수정)")
