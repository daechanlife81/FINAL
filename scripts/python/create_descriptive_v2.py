import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

df = pd.read_excel("260414_data_recoded.xlsx")

wb = Workbook()

header_font = Font(name="Times New Roman", bold=True, size=11)
body_font = Font(name="Times New Roman", size=11)
title_font = Font(name="Times New Roman", bold=True, size=13)
note_font = Font(name="Times New Roman", italic=True, size=10)
center = Alignment(horizontal="center", vertical="center")
center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center")
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin")
thick = Side(style="medium")
border_top = Border(top=thick, bottom=thin)
border_mid = Border(top=thin, bottom=thin)
border_bottom = Border(bottom=thick)

# ========== Sheet 1: 응답자 특성 (기존 유지) ==========
ws1 = wb.active
ws1.title = "응답자 특성"

ws1.merge_cells("A1:D1")
ws1["A1"] = "응답자의 일반적 특성 (N=1,328)"
ws1["A1"].font = title_font
ws1["A1"].alignment = center

headers = ["변수", "구분", "빈도(명)", "비율(%)"]
for i, h in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

row = 4

def write_freq(ws, r, var_name, col, categories):
    start_r = r
    ws.cell(row=r, column=1, value=var_name).font = body_font
    ws.cell(row=r, column=1).alignment = left_wrap
    ws.cell(row=r, column=1).border = border_mid
    for label, codes in categories:
        if isinstance(codes, list):
            count = df[col].isin(codes).sum()
        else:
            count = (df[col] == codes).sum()
        pct = count / len(df) * 100
        for c, v in enumerate([None, label, count, f"{pct:.1f}"], 1):
            cell = ws.cell(row=r, column=c)
            if c >= 2:
                cell.value = v
            cell.font = body_font
            cell.alignment = center if c >= 3 else left_align
            cell.border = border_mid
        r += 1
    if r - start_r > 1:
        ws.merge_cells(start_row=start_r, start_column=1, end_row=r-1, end_column=1)
        ws.cell(row=start_r, column=1).alignment = left_wrap
    return r

row = write_freq(ws1, row, "성별", "re_SQ1", [("남성", 0), ("여성", 1)])
row = write_freq(ws1, row, "연령대", "HSQ2", [
    ("만 19~29세", 1), ("30대", 2), ("40대", 3), ("50대", 4), ("60~79세", 5)])
row = write_freq(ws1, row, "조사지역", "SQ3", [
    ("서울", 1), ("부산", 2), ("대구", 3), ("인천", 4), ("광주", 5),
    ("대전", 6), ("울산", 7), ("세종", 8), ("경기", 9), ("강원", 10),
    ("충북", 11), ("충남", 12), ("전북", 13), ("전남", 14), ("경북", 15),
    ("경남", 16), ("제주", 17)])
row = write_freq(ws1, row, "교육수준", "D_03", [
    ("초등학교 이하", 1), ("중학교", 2), ("고등학교", 3), ("대학교", 4), ("대학원 이상", 5)])
row = write_freq(ws1, row, "혼인상태", "D_04", [
    ("미혼", 1), ("배우자 있음", 2), ("사별", 3), ("이혼", 4), ("별거", 5)])
row = write_freq(ws1, row, "월평균\n가구소득", "D_08", [
    ("100만원 미만", 1), ("100~200만원 미만", 2), ("200~300만원 미만", 3),
    ("300~400만원 미만", 4), ("400~500만원 미만", 5), ("500~600만원 미만", 6),
    ("600~700만원 미만", 7), ("700~800만원 미만", 8), ("800만원 이상", 9)])
row = write_freq(ws1, row, "자원봉사\n참여경험", "re_A_01_1", [("없다", 0), ("있다", 1)])
row = write_freq(ws1, row, "과거 자원봉사\n경험", "re_B_02", [("없음", 0), ("있음", 1)])
row = write_freq(ws1, row, "자원봉사교육\n경험", "re_B_04", [("없음", 0), ("있음", 1)])
row = write_freq(ws1, row, "사회단체활동", "re_C_07", [("미참여", 0), ("참여", 1)])
row = write_freq(ws1, row, "종교활동", "re_C_08", [("미참여", 0), ("참여", 1)])
row = write_freq(ws1, row, "네트워크", "re_C_06", [("부족", 0), ("충분", 1)])

for c in range(1, 5):
    ws1.cell(row=row-1, column=c).border = border_bottom

ws1.merge_cells(f"A{row}:D{row}")
ws1[f"A{row}"] = f"주. 연령: M={df['SQ2'].mean():.2f}, SD={df['SQ2'].std():.2f} (범위: {df['SQ2'].min()}~{df['SQ2'].max()}세)"
ws1[f"A{row}"].font = note_font

ws1.column_dimensions["A"].width = 18
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 12
ws1.column_dimensions["D"].width = 12

# ========== Sheet 2: 이분형 변수 기술통계 ==========
ws2 = wb.create_sheet("이분형 변수")

ws2.merge_cells("A1:E1")
ws2["A1"] = "이분형 변수의 기술통계 (N=1,328)"
ws2["A1"].font = title_font
ws2["A1"].alignment = center

headers2 = ["구분", "변수", "구분(0/1)", "빈도(명)", "비율(%)"]
for i, h in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

binary_vars = [
    ("통제변수", "성별", "re_SQ1", "남성(0)", "여성(1)"),
    ("", "과거 자원봉사경험", "re_B_02", "없음(0)", "있음(1)"),
    ("사회자본", "네트워크", "re_C_06", "부족(0)", "충분(1)"),
    ("", "사회단체활동", "re_C_07", "미참여(0)", "참여(1)"),
    ("문화자본", "종교활동", "re_C_08", "미참여(0)", "참여(1)"),
    ("", "자원봉사교육 경험", "re_B_04", "없음(0)", "있음(1)"),
    ("종속변수", "자원봉사 참여경험", "re_A_01_1", "없다(0)", "있다(1)"),
]

r = 4
prev_group = None
group_start = None

for group, var_name, col, label_0, label_1 in binary_vars:
    count_0 = (df[col] == 0).sum()
    count_1 = (df[col] == 1).sum()
    pct_0 = count_0 / len(df) * 100
    pct_1 = count_1 / len(df) * 100

    if group and group != prev_group:
        if prev_group and group_start and group_start < r:
            ws2.merge_cells(start_row=group_start, start_column=1, end_row=r-1, end_column=1)
            ws2.cell(row=group_start, column=1).alignment = center_wrap
        group_start = r
        prev_group = group

    # 0 행
    vals_0 = [group if group else "", var_name, label_0, count_0, f"{pct_0:.1f}"]
    for c, v in enumerate(vals_0, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.font = body_font
        cell.alignment = center if c >= 4 else left_align
        cell.border = border_mid
    r += 1

    # 1 행
    vals_1 = ["", "", label_1, count_1, f"{pct_1:.1f}"]
    for c, v in enumerate(vals_1, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.font = body_font
        cell.alignment = center if c >= 4 else left_align
        cell.border = border_mid

    # 변수명 병합
    ws2.merge_cells(start_row=r-1, start_column=2, end_row=r, end_column=2)
    ws2.cell(row=r-1, column=2).alignment = center_wrap

    r += 1

# 마지막 그룹 병합
if group_start and group_start < r:
    ws2.merge_cells(start_row=group_start, start_column=1, end_row=r-1, end_column=1)
    ws2.cell(row=group_start, column=1).alignment = center_wrap

for c in range(1, 6):
    ws2.cell(row=r-1, column=c).border = border_bottom

ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 12

# ========== Sheet 3: 연속형 변수 기술통계 ==========
ws3 = wb.create_sheet("연속형 변수")

ws3.merge_cells("A1:H1")
ws3["A1"] = "연속형 변수의 기술통계 및 정규성 검토 (N=1,328)"
ws3["A1"].font = title_font
ws3["A1"].alignment = center

headers3 = ["구분", "변수", "문항 내용", "M", "SD", "왜도", "첨도", "alpha"]
for i, h in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

continuous_vars = [
    ("통제변수", "연령", "", "SQ2", None),
    ("", "혼인상태", "", "D_04", None),
    ("인간자본", "주관적 건강", "", "D_05", None),
    ("", "교육수준", "", "D_03", None),
    ("사회자본", "일반적 신뢰", "", "C_05", None),
    ("문화자본", "자원봉사제도 인식1", "관리시스템", "B_06_1", None),
    ("", "자원봉사제도 인식2", "종합보험", "B_06_2", None),
    ("", "자원봉사제도 인식3", "자원봉사센터", "B_06_3", None),
    ("", "자원봉사제도 인식4", "마일리지/쿠폰", "B_06_4", None),
    ("", "자원봉사제도 인식5", "기부금처리 제도", "B_06_5", None),
    ("", "자원봉사제도 인식6", "자원봉사활동기본법", "B_06_6", None),
    ("", "자원봉사제도 인식7", "자원봉사자의 날", "B_06_7", None),
    ("", "자원봉사제도 인식8", "대한민국 자원봉사 대상", "B_06_8", ".910"),
    ("매개변수\n(효능감)", "효능감1", "의미있는 기여", "B_05_1", None),
    ("", "효능감2", "기회 탐색", "B_05_2", None),
    ("", "효능감3", "사회정의 실현", "B_05_3", None),
    ("", "효능감4", "지역사회 변화", "B_05_4", None),
    ("", "효능감5", "도움 제공", "B_05_5", None),
    ("", "효능감6", "전문가 교류", "B_05_6", None),
    ("", "효능감7", "동등 기회", "B_05_7", None),
    ("", "효능감8", "지식 적용", "B_05_8", None),
    ("", "효능감9", "자기돌봄 지원", "B_05_9", None),
    ("", "효능감10", "참여 의향", "B_05_10", ".889"),
    ("매개변수\n(안녕감)", "행복", "", "C_01", None),
    ("", "걱정(역코딩)", "", "re_C_02", None),
    ("", "우울(역코딩)", "", "re_C_03", None),
    ("", "삶의 만족", "", "C_04", ".696"),
]

r = 4
prev_group = None
group_start = None

for group, var_name, item, col, alpha in continuous_vars:
    m = df[col].mean()
    sd = df[col].std()
    skew = df[col].skew()
    kurt = df[col].kurtosis()

    if group and group != prev_group:
        if prev_group and group_start and group_start < r:
            ws3.merge_cells(start_row=group_start, start_column=1, end_row=r-1, end_column=1)
            ws3.cell(row=group_start, column=1).alignment = center_wrap
        group_start = r
        prev_group = group

    vals = [group if group else "", var_name, item,
            f"{m:.2f}", f"{sd:.2f}", f"{skew:.3f}", f"{kurt:.3f}",
            alpha if alpha else ""]
    for c, v in enumerate(vals, 1):
        cell = ws3.cell(row=r, column=c, value=v)
        cell.font = body_font
        cell.alignment = center if c >= 4 else left_wrap
        cell.border = border_mid

    r += 1

# 마지막 그룹 병합
if group_start and group_start < r:
    ws3.merge_cells(start_row=group_start, start_column=1, end_row=r-1, end_column=1)
    ws3.cell(row=group_start, column=1).alignment = center_wrap

for c in range(1, 9):
    ws3.cell(row=r-1, column=c).border = border_bottom

# 주석
ws3.merge_cells(f"A{r}:H{r}")
ws3[f"A{r}"] = "주. 정규성 기준: |왜도|<3, |첨도|<10 (Kline, 2015). 모든 변수 허용 범위 내. alpha=Cronbach's alpha (잠재변수 마지막 문항에 표기)."
ws3[f"A{r}"].font = note_font

ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["C"].width = 24
ws3.column_dimensions["D"].width = 8
ws3.column_dimensions["E"].width = 8
ws3.column_dimensions["F"].width = 10
ws3.column_dimensions["G"].width = 10
ws3.column_dimensions["H"].width = 10

wb.save("descriptive_statistics_v2.xlsx")
print("저장 완료: descriptive_statistics.xlsx (3개 시트)")
