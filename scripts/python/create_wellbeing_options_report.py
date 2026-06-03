"""
주관적 안녕감 측정 옵션 3가지 비교 종합 보고서 엑셀 생성
박사학위논문 마지막 점검 - 지도교수님 보고용

방안 1: 4문항 (현재) - 행복+만족+우울역+걱정역
방안 2: 2문항 (지도교수님 권장) - 행복+만족
방안 3: 분리 (학술적 최우수) - 양적(행복+만족) + 음적(우울역+걱정역)

파일명: Wellbeing_Measurement_Options_Report.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = Workbook()
wb.remove(wb.active)

# 스타일
title_font = Font(name="Times New Roman", bold=True, size=14)
section_font = Font(name="Times New Roman", bold=True, size=12)
header_font = Font(name="Times New Roman", bold=True, size=11)
body_font = Font(name="Times New Roman", size=11)
bold_font = Font(name="Times New Roman", bold=True, size=11)
note_font = Font(name="Times New Roman", italic=True, size=10)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin")
thick = Side(style="medium")
border_top = Border(top=thick, bottom=thin)
border_mid = Border(top=thin, bottom=thin)
border_bottom = Border(bottom=thick)
header_fill = PatternFill("solid", fgColor="D9E1F2")
section_fill = PatternFill("solid", fgColor="F2F2F2")
recommend_fill = PatternFill("solid", fgColor="E2EFDA")


def add_title(ws, text, ncol):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = title_font
    cell.alignment = center
    ws.row_dimensions[1].height = 26


def write_header(ws, headers, row=3, fill=header_fill):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.border = border_top
        cell.fill = fill


def write_data(ws, rows, start_row=4, num_cols=None, bold_first=False, fills=None):
    last = start_row
    for r, row_data in enumerate(rows, start_row):
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = bold_font if (bold_first and c == 1) else body_font
            cell.alignment = center if c > 1 else left
            cell.border = border_mid
            if fills and (r - start_row) in fills:
                cell.fill = fills[r - start_row]
        last = r
    if num_cols is None:
        num_cols = len(rows[0])
    for c in range(1, num_cols + 1):
        ws.cell(row=last, column=c).border = border_bottom
    return last


def add_note(ws, note, start_row, ncol):
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=ncol)
    cell = ws.cell(row=start_row, column=1, value=note)
    cell.font = note_font
    cell.alignment = left


# ============================================================
# Sheet 1: Executive Summary (지도교수님 보고용 요약)
# ============================================================
ws1 = wb.create_sheet("1. Executive Summary")
add_title(ws1, "주관적 안녕감 측정 옵션 비교 종합 보고서 (Executive Summary)", 5)

# 옵션 비교 표
write_header(ws1,
             ["구분", "방안 1 (현재)", "방안 2 (교수님 권장)", "방안 3 (학술 최우수)",
              "비고"], row=3)

summary = [
    ["측정 문항", "4문항 (만족+행복+우울역+걱정역)",
     "2문항 (만족+행복)", "분리 (양적 + 음적)", "방안 2/3 모두 우울·걱정 별도 처리"],
    ["잠재변수 수", "1개 (wellbeing)", "1개 (wellbeing)", "2개 (pos_wb, neg_wb)", "복잡도 증가"],
    ["Cronbach's α", ".696", ".767", "양:.767, 음:.74 추정", "방안 2/3 모두 .70+"],
    ["AVE", ".388", ".649", "양:.633, 음:.638", "방안 2/3 모두 .50+"],
    ["CR", ".701", ".784", "양:.774, 음:.773", "방안 2/3 모두 .70+"],
    ["CFA CFI", ".688 ❌", ".952 ✅", ".947 ✅", "방안 1만 미달"],
    ["CFA RMSEA", ".406 ❌", ".052 ✅", ".051 ✅", "방안 1만 미달"],
    ["SEM CFI", ".882 △", ".919 ✅", ".911 ✅", "방안 2/3 우수"],
    ["SEM RMSEA", ".054 ✅", ".046 ✅", ".047 ✅", "모두 충족"],
    ["효능감→안녕감", "β=.294***", "β=.301***", "양:β=.309***, 음:β=.173***", "모두 유의"],
    ["안녕감→참여", "β=.083* 유의", "β=.073 (p=.050) 한계", "양:β=.065 n.s., 음:β=.021 n.s.", "방안 1만 명확히 유의"],
    ["순차매개 유의 변수", "5개", "3개 (건강·제도·봉사교육)", "0개 (음수분산 경고)", "방안 1 가장 풍부"],
    ["학술적 권위", "Diener 3요소", "World Happiness Report 표준", "Diener 2차원 분리", "각각 근거 있음"],
    ["분석 안정성", "안정", "안정", "음수 분산 경고", "방안 3 주의"],
    ["권장도", "★★★", "★★★★", "★★★", "방안 2 가장 권장"],
]

# Recommend row (방안 2)
fills = {len(summary)-1: recommend_fill}
last = write_data(ws1, summary, start_row=4, num_cols=5, bold_first=True)

add_note(ws1, "주. 방안 2 (지도교수님 권장)가 학술적·실용적 균형에서 가장 우수",
         last + 2, 5)
add_note(ws1, "    방안 3은 양/음 분리로 학술적 정교성 ↑ 그러나 음수 분산 경고 발생",
         last + 3, 5)
add_note(ws1, "    14개 가설 중 안녕감→참여(H12), 안녕감 매개(H14) 결과 변동 발생",
         last + 4, 5)

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 26
ws1.column_dimensions["D"].width = 30
ws1.column_dimensions["E"].width = 30


# ============================================================
# Sheet 2: CFA 적합도 비교
# ============================================================
ws2 = wb.create_sheet("2. CFA Comparison")
add_title(ws2, "측정모형 (CFA) 적합도 비교", 5)

write_header(ws2,
             ["적합도 지수", "방안 1 (현재)", "방안 2 (2문항)", "방안 3 (분리)",
              "기준"], row=3)

cfa_data = [
    ["χ²", "1353.408", "762.764", "903.663", "-"],
    ["df", "206", "167", "203", "-"],
    ["χ²/df", "6.570", "4.567", "4.452", "≤5 수용"],
    ["GFI", ".913 ✅", ".943 ✅", ".938 ✅", "≥.90"],
    ["CFI", ".913 ✅", ".952 ✅", ".947 ✅", "≥.90"],
    ["TLI", ".903 ✅", ".945 ✅", ".940 ✅", "≥.90"],
    ["RMSEA", ".065 ✅", ".052 ✅", ".051 ✅", "≤.08"],
    ["SRMR", ".051 ✅", ".039 ✅", ".042 ✅", "≤.08"],
]
last = write_data(ws2, cfa_data, start_row=4, num_cols=5, bold_first=True)

add_note(ws2, "주. 측정모형 적합도는 모든 방안이 충족, 방안 2가 가장 우수",
         last + 2, 5)

ws2.column_dimensions["A"].width = 14
for col in "BCDE":
    ws2.column_dimensions[col].width = 16


# ============================================================
# Sheet 3: SEM 적합도 비교
# ============================================================
ws3 = wb.create_sheet("3. SEM Comparison")
add_title(ws3, "구조모형 (SEM) 적합도 비교 (MLR 추정)", 5)

write_header(ws3,
             ["적합도 지수", "방안 1 (현재)", "방안 2 (2문항)", "방안 3 (분리)",
              "기준"], row=3)

sem_data = [
    ["χ² (S-B)", "2092.837", "1442.742", "1678.654", "-"],
    ["df", "447", "384", "435", "-"],
    ["χ²/df", "4.682", "3.757", "3.859", "≤5 수용"],
    ["GFI", ".872", ".900 ✅", ".894", "≥.90"],
    ["CFI (robust)", ".882", ".919 ✅", ".911 ✅", "≥.90"],
    ["TLI (robust)", ".866", ".907 ✅", ".896", "≥.90"],
    ["RMSEA (robust)", ".054 ✅", ".046 ✅", ".047 ✅", "≤.06"],
    ["SRMR", ".074 ✅", ".073 ✅", ".073 ✅", "≤.08"],
]
last = write_data(ws3, sem_data, start_row=4, num_cols=5, bold_first=True)

add_note(ws3, "주. SEM 적합도: 방안 2가 가장 우수 (CFI, TLI 모두 .90 이상)",
         last + 2, 5)
add_note(ws3, "    방안 1은 CFI=.882로 .90 미달, 방안 3은 음수 분산 경고",
         last + 3, 5)

ws3.column_dimensions["A"].width = 16
for col in "BCDE":
    ws3.column_dimensions[col].width = 16


# ============================================================
# Sheet 4: 가설 검증 결과 비교
# ============================================================
ws4 = wb.create_sheet("4. Hypothesis Comparison")
add_title(ws4, "14개 가설 검증 결과 비교", 4)

write_header(ws4, ["가설", "내용", "방안 1 (현재)", "방안 2 (2문항)"], row=3)

hyp_data = [
    ["H1", "인간자본 → 효능감", "채택", "채택"],
    ["H2", "사회자본 → 효능감", "부분채택", "부분채택"],
    ["H3", "문화자본 → 효능감", "채택", "채택"],
    ["H4", "인간자본 → 안녕감", "부분채택 (건강)", "부분채택 (건강)"],
    ["H5", "사회자본 → 안녕감", "부분채택 (신뢰·네트워크)", "부분채택 (신뢰만)"],
    ["H6", "종교활동 → 안녕감", "기각", "기각"],
    ["H7", "효능감 → 안녕감", "채택 (β=.294***)", "채택 (β=.301***)"],
    ["H8", "인간자본 → 참여", "기각", "기각"],
    ["H9", "사회자본 → 참여", "부분채택 (사회단체)", "부분채택 (사회단체)"],
    ["H10", "문화자본 → 참여", "채택", "채택"],
    ["H11", "효능감 → 참여", "기각", "기각"],
    ["H12", "안녕감 → 참여", "채택 (β=.083*)", "한계 (β=.073, p=.050)"],
    ["H13", "효능감 단순매개", "기각", "기각"],
    ["H14", "안녕감 매개 (인간·사회·종교)", "부분채택 (3개 유의)", "부분채택 (2개 유의: 건강·신뢰)"],
]
last = write_data(ws4, hyp_data, start_row=4, num_cols=4, bold_first=True)

# 종합
summary_row = last + 1
ws4.merge_cells(start_row=summary_row, start_column=1,
                end_row=summary_row, end_column=4)
ws4.cell(row=summary_row, column=1,
         value="종합: 방안 1 채택 5/부분채택 5/기각 4 | 방안 2 채택 5/부분채택 4/기각 5").font = bold_font
ws4.cell(row=summary_row, column=1).alignment = center
ws4.cell(row=summary_row, column=1).fill = section_fill

add_note(ws4, "주. H12, H14에서 결과 변동 발생 (안녕감 측정 차이로 인함)",
         summary_row + 2, 4)
add_note(ws4, "    방안 2에서 H12가 p=.050으로 경계선상에 위치",
         summary_row + 3, 4)
add_note(ws4, "    방안 3은 모형 복잡성으로 별도 비교표 미작성 (Sheet 5 참조)",
         summary_row + 4, 4)

ws4.column_dimensions["A"].width = 8
ws4.column_dimensions["B"].width = 36
ws4.column_dimensions["C"].width = 30
ws4.column_dimensions["D"].width = 30


# ============================================================
# Sheet 5: 방안 3 (양적/음적 분리) 결과
# ============================================================
ws5 = wb.create_sheet("5. Option 3 Results")
add_title(ws5, "방안 3 (양적/음적 안녕감 분리) 상세 결과", 4)

# 양적 안녕감 경로
ws5.cell(row=3, column=1, value="[1] 양적 안녕감 (pos_wb) 영향요인").font = section_font
ws5.cell(row=3, column=1).fill = section_fill
ws5.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)

write_header(ws5, ["변수", "β", "p", "유의"], row=4)
pos_wb_data = [
    ["주관적 건강", .208, "<.001", "***"],
    ["일반적 신뢰", .282, "<.001", "***"],
    ["교육수준", .055, ".156", "n.s."],
    ["네트워크", .038, ".190", "n.s."],
    ["사회단체활동", .000, ".998", "n.s."],
    ["종교활동", -.010, ".751", "n.s."],
    ["효능감", .309, "<.001", "***"],
]
last = write_data(ws5, pos_wb_data, start_row=5, num_cols=4, bold_first=True)

# 음적 안녕감 경로
ws5.cell(row=last+2, column=1,
         value="[2] 음적 안녕감 (neg_wb, 우울·걱정 부재) 영향요인").font = section_font
ws5.cell(row=last+2, column=1).fill = section_fill
ws5.merge_cells(start_row=last+2, start_column=1,
                end_row=last+2, end_column=4)

write_header(ws5, ["변수", "β", "p", "유의"], row=last+3)
neg_wb_data = [
    ["주관적 건강", .116, ".001", "***"],
    ["네트워크", .156, "<.001", "***"],
    ["사회단체활동", .105, "<.001", "***"],
    ["종교활동", -.071, ".004", "**"],
    ["일반적 신뢰", .045, ".080", "n.s."],
    ["교육수준", -.042, ".220", "n.s."],
    ["효능감", .173, "<.001", "***"],
]
last2 = write_data(ws5, neg_wb_data, start_row=last+4, num_cols=4, bold_first=True)

# 참여 경로
ws5.cell(row=last2+2, column=1,
         value="[3] 자원봉사 참여 영향요인 (매개변수)").font = section_font
ws5.cell(row=last2+2, column=1).fill = section_fill
ws5.merge_cells(start_row=last2+2, start_column=1,
                end_row=last2+2, end_column=4)

write_header(ws5, ["변수", "β", "p", "유의"], row=last2+3)
parti_data = [
    ["양적 안녕감 → 참여", .065, ".104", "n.s."],
    ["음적 안녕감 → 참여", .021, ".431", "n.s."],
    ["효능감 → 참여", .006, ".841", "n.s."],
]
last3 = write_data(ws5, parti_data, start_row=last2+4, num_cols=4, bold_first=True)

add_note(ws5,
         "주요 발견: 양적 안녕감은 신뢰·건강·효능감에 강한 영향, 음적 안녕감은 네트워크·사회단체·효능감에 영향",
         last3 + 2, 4)
add_note(ws5,
         "        효능감은 양적 안녕감 (β=.309)에 더 강한 영향, 음적 안녕감 (β=.173)에는 약한 영향",
         last3 + 3, 4)
add_note(ws5,
         "        그러나 양적/음적 안녕감 모두 자원봉사 참여에 유의하지 않음 (β=.065, .021)",
         last3 + 4, 4)
add_note(ws5,
         "        한계: 모형 추정 과정에서 음수 분산 경고 발생, 부트스트랩 4412회 nonadmissible",
         last3 + 5, 4)

ws5.column_dimensions["A"].width = 22
ws5.column_dimensions["B"].width = 14
ws5.column_dimensions["C"].width = 14
ws5.column_dimensions["D"].width = 12


# ============================================================
# Sheet 6: 매개효과 비교
# ============================================================
ws6 = wb.create_sheet("6. Mediation Comparison")
add_title(ws6, "안녕감 매개효과 비교 (95% Bootstrap CI)", 7)

write_header(ws6,
             ["독립변수", "방안 1 (현재) β", "방안 1 유의",
              "방안 2 (2문항) β", "방안 2 유의",
              "방안 3 양적 β", "방안 3 유의"],
             row=3)

med_compare = [
    ["주관적 건강", 0.0192, "유의", 0.0152, "유의", 0.0135, "비유의"],
    ["교육수준", 0.0030, "비유의", 0.0040, "비유의", 0.0036, "비유의"],
    ["일반적 신뢰", 0.0224, "유의", 0.0205, "유의", 0.0183, "비유의"],
    ["네트워크", 0.0063, "유의", 0.0028, "비유의", 0.0025, "비유의"],
    ["사회단체활동", 0.0030, "비유의", 0.0000, "비유의", 0.0000, "비유의"],
    ["종교활동", -0.0019, "비유의", -0.0006, "비유의", -0.0006, "비유의"],
]
last = write_data(ws6, med_compare, start_row=4, num_cols=7, bold_first=True)

add_note(ws6, "주. 단순매개: 독립 → 안녕감 → 참여 경로",
         last + 2, 7)
add_note(ws6, "    방안 1: 3개 유의 (건강, 신뢰, 네트워크)",
         last + 3, 7)
add_note(ws6, "    방안 2: 2개 유의 (건강, 신뢰)",
         last + 4, 7)
add_note(ws6, "    방안 3: 양적 안녕감 매개 0개 유의 (음수 분산 문제로 신뢰성 낮음)",
         last + 5, 7)

ws6.column_dimensions["A"].width = 16
for col in "BCDEFG":
    ws6.column_dimensions[col].width = 16


# ============================================================
# Sheet 7: 학술적 권장 사항
# ============================================================
ws7 = wb.create_sheet("7. Recommendation")
add_title(ws7, "학술적 권장 사항 (지도교수님 검토용)", 2)

# 권장 사항
ws7.cell(row=3, column=1, value="구분").font = header_font
ws7.cell(row=3, column=2, value="내용").font = header_font
ws7.cell(row=3, column=1).fill = header_fill
ws7.cell(row=3, column=2).fill = header_fill
ws7.cell(row=3, column=1).border = border_top
ws7.cell(row=3, column=2).border = border_top
ws7.cell(row=3, column=1).alignment = center
ws7.cell(row=3, column=2).alignment = center

rec_data = [
    ["1. 방안 1 (현재)",
     "장점: 현재까지 작성한 모든 분석 유지 / Diener 3요소 일관성\n"
     "단점: CFA 적합도 미달 (CFI=.688), AVE 부족 (.388)\n"
     "권장도: ★★★"],
    ["2. 방안 2 (교수님 권장)",
     "장점: 신뢰도 향상 (α=.767), AVE/CR 충족, CFI .919\n"
     "      국제 표준(World Happiness Report)과 부합\n"
     "단점: H12(안녕감→참여) p=.050 경계선, 4문항 → 2문항 축소\n"
     "권장도: ★★★★ (강력 권장)"],
    ["3. 방안 3 (학술 최우수)",
     "장점: 양적/음적 안녕감의 차별적 영향 검증 가능\n"
     "      양적 안녕감은 신뢰·건강 중심, 음적 안녕감은 네트워크·사회단체 중심\n"
     "단점: 음수 분산 경고, 부트스트랩 4412회 nonadmissible, 안녕감 매개 모두 비유의\n"
     "      모형 안정성 우려\n"
     "권장도: ★★★ (학술 가치 있으나 모형 불안정)"],
    ["최종 추천", "방안 2를 본 분석으로 채택하고, 방안 3을 부록에 보조 분석으로 제시\n"
     "이는 (1) 지도교수님 권장 반영, (2) 학술적 신뢰도 확보,\n"
     "(3) 양적/음적 차별 검증의 보조 근거 제공이라는 3중 강점을 갖춤"],
]

start = 4
for r, (label, content) in enumerate(rec_data, start):
    cell1 = ws7.cell(row=r, column=1, value=label)
    cell1.font = bold_font
    cell1.alignment = center
    cell1.border = border_mid
    cell2 = ws7.cell(row=r, column=2, value=content)
    cell2.font = body_font
    cell2.alignment = left
    cell2.border = border_mid
    ws7.row_dimensions[r].height = 80
    if r == start + 3:  # 최종 추천
        cell1.fill = recommend_fill
        cell2.fill = recommend_fill

for col in range(1, 3):
    ws7.cell(row=start + len(rec_data) - 1, column=col).border = border_bottom

ws7.column_dimensions["A"].width = 22
ws7.column_dimensions["B"].width = 90


# 저장
wb.save("results/Wellbeing_Measurement_Options_Report.xlsx")
print("저장 완료: results/Wellbeing_Measurement_Options_Report.xlsx")
print("총 7개 시트:")
print("  1. Executive Summary (지도교수님 보고용 요약)")
print("  2. CFA Comparison (측정모형 적합도 비교)")
print("  3. SEM Comparison (구조모형 적합도 비교)")
print("  4. Hypothesis Comparison (14개 가설 비교)")
print("  5. Option 3 Results (방안 3 상세 결과)")
print("  6. Mediation Comparison (매개효과 비교)")
print("  7. Recommendation (학술적 권장 사항)")
