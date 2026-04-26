from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

wb = Workbook()

header_font = Font(name="Times New Roman", bold=True, size=11)
body_font = Font(name="Times New Roman", size=11)
title_font = Font(name="Times New Roman", bold=True, size=13)
note_font = Font(name="Times New Roman", italic=True, size=10)
center = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin")
thick = Side(style="medium")
border_top = Border(top=thick, bottom=thin)
border_mid = Border(top=thin, bottom=thin)
border_bottom = Border(bottom=thick)

# ========== Sheet 1 ==========
ws1 = wb.active
ws1.title = "측정모형 적합도"

ws1.merge_cells("A1:C1")
ws1["A1"] = "측정모형 적합도 검증 결과"
ws1["A1"].font = title_font
ws1["A1"].alignment = center

for i, h in enumerate(["적합도 지수", "값", "기준"], 1):
    cell = ws1.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

fit_data = [
    ("χ²(df)", "1353.408(206), p<.001", "-"),
    ("χ²/df", "6.570", "≤3 양호, ≤5 수용"),
    ("GFI", ".913", "≥.90"),
    ("CFI", ".913", "≥.90"),
    ("TLI", ".903", "≥.90"),
    ("RMSEA", ".065 [.062, .068]", "≤.08"),
    ("SRMR", ".051", "≤.08"),
]

for r, (a, b, c) in enumerate(fit_data, 4):
    for col, v in enumerate([a, b, c], 1):
        cell = ws1.cell(row=r, column=col, value=v)
        cell.font = body_font
        cell.alignment = center
        cell.border = border_mid

for c in range(1, 4):
    ws1.cell(row=10, column=c).border = border_bottom

ws1.merge_cells("A11:C11")
ws1["A11"] = "주. χ²/df가 다소 높으나 대표본(N=1,328) 효과에 기인함(Bentler & Bonett, 1980)."
ws1["A11"].font = note_font

ws1.column_dimensions["A"].width = 16
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 28

# ========== Sheet 2 ==========
ws2 = wb.create_sheet("확인적 요인분석")

ws2.merge_cells("A1:F1")
ws2["A1"] = "확인적 요인분석 결과"
ws2["A1"].font = title_font
ws2["A1"].alignment = center

headers2 = ["잠재변수", "관측변수", "B", "S.E.", "β", "p"]
for i, h in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

cfa_data = [
    ("자원봉사 효능감", "B_05_1", "1.000", "-", ".685", "<.001"),
    ("", "B_05_2", "1.047", ".048", ".619", "<.001"),
    ("", "B_05_3", "1.173", ".048", ".685", "<.001"),
    ("", "B_05_4", "1.189", ".051", ".654", "<.001"),
    ("", "B_05_5", "1.106", ".047", ".662", "<.001"),
    ("", "B_05_6", "1.076", ".049", ".622", "<.001"),
    ("", "B_05_7", "1.264", ".049", ".712", "<.001"),
    ("", "B_05_8", "1.170", ".050", ".660", "<.001"),
    ("", "B_05_9", "1.142", ".046", ".692", "<.001"),
    ("", "B_05_10", "1.114", ".045", ".693", "<.001"),
    ("자원봉사제도 인식", "B_06_1", "1.000", "-", ".734", "<.001"),
    ("", "B_06_2", "0.923", ".030", ".781", "<.001"),
    ("", "B_06_3", "0.855", ".033", ".627", "<.001"),
    ("", "B_06_4", "1.087", ".029", ".805", "<.001"),
    ("", "B_06_5", "1.046", ".033", ".735", "<.001"),
    ("", "B_06_6", "1.036", ".027", ".815", "<.001"),
    ("", "B_06_7", "0.963", ".028", ".774", "<.001"),
    ("", "B_06_8", "1.003", ".033", ".729", "<.001"),
    ("주관적 안녕감", "C_01", "1.000", "-", ".799", "<.001"),
    ("", "re_C_02(역)", "0.771", ".056", ".375", "<.001"),
    ("", "re_C_03(역)", "0.911", ".051", ".457", "<.001"),
    ("", "C_04", "0.992", ".041", ".748", "<.001"),
]

for r, (latent, obs, b, se, beta, p) in enumerate(cfa_data, 4):
    for col, v in enumerate([latent, obs, b, se, beta, p], 1):
        cell = ws2.cell(row=r, column=col, value=v)
        cell.font = body_font
        cell.alignment = center
        cell.border = border_mid

for c in range(1, 7):
    ws2.cell(row=25, column=c).border = border_bottom

ws2.merge_cells("A4:A13")
ws2.merge_cells("A14:A21")
ws2.merge_cells("A22:A25")
for r in [4, 14, 22]:
    ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 10
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 10

# ========== Sheet 3 ==========
ws3 = wb.create_sheet("타당도 및 신뢰도")

ws3.merge_cells("A1:E1")
ws3["A1"] = "집중타당도, 판별타당도 및 신뢰도 검증 결과"
ws3["A1"].font = title_font
ws3["A1"].alignment = center

ws3.merge_cells("A3:E3")
ws3["A3"] = "1. 집중타당도 및 신뢰도"
ws3["A3"].font = Font(name="Times New Roman", bold=True, size=11)

for i, h in enumerate(["잠재변수", "AVE", "CR", "Cronbach's α", "판단"], 1):
    cell = ws3.cell(row=4, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

conv_data = [
    ("자원봉사 효능감", ".447", ".890", ".889", "CR 충족"),
    ("자원봉사제도 인식", ".566", ".912", ".910", "모두 충족"),
    ("주관적 안녕감", ".387", ".698", ".696", "CR 기준 수용*"),
]
for r, row_data in enumerate(conv_data, 5):
    for c, v in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=v)
        cell.font = body_font
        cell.alignment = center
        cell.border = border_mid

for c in range(1, 6):
    ws3.cell(row=7, column=c).border = border_bottom

ws3.merge_cells("A8:E8")
ws3["A8"] = "주. 기준: AVE≥.50, CR≥.70, α≥.70. *CR≥.60이면 AVE<.50도 수용 가능(Fornell & Larcker, 1981)."
ws3["A8"].font = note_font

ws3.merge_cells("A10:D10")
ws3["A10"] = "2. 판별타당도 (Fornell-Larcker 기준)"
ws3["A10"].font = Font(name="Times New Roman", bold=True, size=11)

for i, h in enumerate(["", "자원봉사 효능감", "자원봉사제도 인식", "주관적 안녕감"], 1):
    cell = ws3.cell(row=11, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

discrim = [
    ("자원봉사 효능감", "(.447)", "", ""),
    ("자원봉사제도 인식", ".450 (.202)", "(.566)", ""),
    ("주관적 안녕감", ".407 (.166)", ".269 (.072)", "(.387)"),
]
for r, row_data in enumerate(discrim, 12):
    for c, v in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=v)
        cell.font = body_font
        cell.alignment = center
        cell.border = border_mid

for c in range(1, 5):
    ws3.cell(row=14, column=c).border = border_bottom

ws3.merge_cells("A15:D15")
ws3["A15"] = "주. 대각선: (AVE), 하삼각: 상관계수 (상관제곱). 판별타당도 조건: AVE > r². 모두 충족."
ws3["A15"].font = note_font

ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 20
ws3.column_dimensions["C"].width = 20
ws3.column_dimensions["D"].width = 18
ws3.column_dimensions["E"].width = 14

wb.save("측정모형_타당도_신뢰도.xlsx")
print("저장 완료: 측정모형_타당도_신뢰도.xlsx")
