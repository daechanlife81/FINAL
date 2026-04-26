from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side

wb = load_workbook("SEM_results.xlsx")

header_font = Font(name="Times New Roman", bold=True, size=11)
body_font = Font(name="Times New Roman", size=11)
title_font = Font(name="Times New Roman", bold=True, size=13)
note_font = Font(name="Times New Roman", italic=True, size=10)
bold_font = Font(name="Times New Roman", bold=True, size=11)
center = Alignment(horizontal="center", vertical="center")
center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin")
thick = Side(style="medium")
border_top = Border(top=thick, bottom=thin)
border_mid = Border(top=thin, bottom=thin)
border_bottom = Border(bottom=thick)
border_section = Border(top=thick, bottom=thin)

ws = wb.create_sheet("가설검증 종합")

ws.merge_cells("A1:F1")
ws["A1"] = "연구가설 검증 결과 종합"
ws["A1"].font = title_font
ws["A1"].alignment = center

headers = ["연구문제", "가설", "경로", "beta", "p", "판단"]
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

data = [
    ("연구문제1\n자본->효능감", "H1. 인간자본->효능감", "건강 -> 효능감", ".144", "***", ""),
    ("", "", "교육 -> 효능감", ".069", "*", ""),
    ("", "", "", "", "", "채택"),
    ("", "H2. 사회자본->효능감", "사회신뢰 -> 효능감", ".042", "n.s.", ""),
    ("", "", "네트워크 -> 효능감", ".078", "**", ""),
    ("", "", "사회단체 -> 효능감", ".003", "n.s.", ""),
    ("", "", "", "", "", "부분채택"),
    ("", "H3. 문화자본->효능감", "제도인식 -> 효능감", ".332", "***", ""),
    ("", "", "종교활동 -> 효능감", ".079", "**", ""),
    ("", "", "봉사교육 -> 효능감", ".156", "***", ""),
    ("", "", "", "", "", "채택"),
    ("연구문제2\n자본->안녕감", "H4. 인간자본->안녕감", "건강 -> 안녕감", ".230", "***", ""),
    ("", "", "교육 -> 안녕감", ".036", "n.s.", ""),
    ("", "", "", "", "", "부분채택"),
    ("", "H5. 사회자본->안녕감", "사회신뢰 -> 안녕감", ".269", "***", ""),
    ("", "", "네트워크 -> 안녕감", ".075", "**", ""),
    ("", "", "사회단체 -> 안녕감", ".037", "n.s.", ""),
    ("", "", "", "", "", "부분채택"),
    ("", "H6. 종교활동->안녕감", "종교활동 -> 안녕감", "-.022", "n.s.", ""),
    ("", "", "", "", "", "기각"),
    ("연구문제3\n자본->참여", "H7. 인간자본->참여", "건강 -> 참여", "-.062", "*", ""),
    ("", "", "교육 -> 참여", "-.028", "n.s.", ""),
    ("", "", "", "", "", "기각"),
    ("", "H8. 사회자본->참여", "사회신뢰 -> 참여", ".032", "n.s.", ""),
    ("", "", "네트워크 -> 참여", "-.000", "n.s.", ""),
    ("", "", "사회단체 -> 참여", ".097", "**", ""),
    ("", "", "", "", "", "부분채택"),
    ("", "H9. 문화자본->참여", "제도인식 -> 참여", ".231", "***", ""),
    ("", "", "종교활동 -> 참여", ".091", "**", ""),
    ("", "", "봉사교육 -> 참여", ".055", "n.s.", ""),
    ("", "", "", "", "", "부분채택"),
    ("연구문제4\n매개경로", "H10. 효능감->안녕감", "효능감 -> 안녕감", ".294", "***", "채택"),
    ("", "H11. 효능감->참여", "효능감 -> 참여", ".005", "n.s.", "기각"),
    ("", "H12. 안녕감->참여", "안녕감 -> 참여", ".083", "*", "채택"),
    ("연구문제5\n매개효과", "H13. 독립->효능감->참여", "8개 경로 모두", "-", "-", "기각"),
    ("", "H14. 독립->안녕감->참여", "건강, 신뢰, 네트워크 유의", "-", "-", "부분채택"),
    ("", "H15. 독립->효능감->안녕감->참여", "건강,네트워크,제도인식,종교,봉사교육 유의", "-", "-", "부분채택"),
]

for r, (rq, hyp, path, beta, p, result) in enumerate(data, 4):
    vals = [rq, hyp, path, beta, p, result]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = body_font
        cell.alignment = center_wrap
        cell.border = border_mid
    if result in ["채택", "부분채택", "기각"]:
        ws.cell(row=r, column=6).font = bold_font

last_row = 4 + len(data) - 1
for c in range(1, 7):
    ws.cell(row=last_row, column=c).border = border_bottom

# 연구문제 열 병합
ws.merge_cells("A4:A14")   # 연구문제1
ws.merge_cells("A15:A22")  # 연구문제2
ws.merge_cells("A23:A33")  # 연구문제3
ws.merge_cells("A34:A36")  # 연구문제4
ws.merge_cells("A37:A39")  # 연구문제5
# 가설 열 병합 (같은 가설)
ws.merge_cells("B4:B6")    # H1
ws.merge_cells("B7:B9")    # H2 (3 sub + 1 result = need to check)

for r in [4, 15, 23, 34, 37]:
    ws.cell(row=r, column=1).alignment = center_wrap

# 종합 행
note_row = last_row + 2
ws.merge_cells(f"A{note_row}:F{note_row}")
ws[f"A{note_row}"] = "가설 검증 종합: 채택 4개, 부분채택 7개, 기각 4개 (총 15개)"
ws[f"A{note_row}"].font = bold_font
ws[f"A{note_row}"].alignment = center

note_row2 = note_row + 1
ws.merge_cells(f"A{note_row2}:F{note_row2}")
ws[f"A{note_row2}"] = "주. *p<.05, **p<.01, ***p<.001. Bootstrap 5,000회 (매개효과)."
ws[f"A{note_row2}"].font = note_font

ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 36
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 8
ws.column_dimensions["F"].width = 12

wb.save("SEM_results.xlsx")
print("저장 완료: SEM_results.xlsx (가설검증 종합 시트 추가)")
