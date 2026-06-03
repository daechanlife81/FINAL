"""
방안 2 (안녕감 2문항) 측정모형 CFA 결과 엑셀 생성
박사학위논문 3.1. 확인적 요인분석 갱신용

파일명: CFA_option2_measurement.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = Workbook()
wb.remove(wb.active)

title_font = Font(name="Times New Roman", bold=True, size=13)
header_font = Font(name="Times New Roman", bold=True, size=11)
body_font = Font(name="Times New Roman", size=11)
bold_font = Font(name="Times New Roman", bold=True, size=11)
note_font = Font(name="Times New Roman", italic=True, size=10)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin")
thick = Side(style="medium")
border_top = Border(top=thick, bottom=thin)
border_mid = Border(top=thin, bottom=thin)
border_bottom = Border(bottom=thick)
header_fill = PatternFill("solid", fgColor="E8E8E8")
hl_fill = PatternFill("solid", fgColor="FFF2CC")


def add_title(ws, text, ncol):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1, value=text)
    c.font = title_font
    c.alignment = center
    ws.row_dimensions[1].height = 22


def write_header(ws, headers, row=3):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.border = border_top
        cell.fill = header_fill


def write_data(ws, rows, start_row=4, num_cols=None, bold_first=False, hl_rows=None):
    last = start_row
    hl_rows = hl_rows or set()
    for r, row_data in enumerate(rows, start_row):
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = bold_font if (bold_first and c == 1) else body_font
            cell.alignment = center if c > 1 else left
            cell.border = border_mid
            if (r - start_row) in hl_rows:
                cell.fill = hl_fill
        last = r
    if num_cols is None:
        num_cols = len(rows[0])
    for c in range(1, num_cols + 1):
        ws.cell(row=last, column=c).border = border_bottom
    return last


def add_note(ws, note, start_row, ncol):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncol)
    c = ws.cell(row=start_row, column=1, value=note)
    c.font = note_font
    c.alignment = left


# ============================================================
# Sheet 1: 측정모형 적합도 비교 (표 9)
# ============================================================
ws1 = wb.create_sheet("1. Model Fit (Table 9)")
add_title(ws1, "측정모형 적합도 비교: 방안 1(4문항) vs 방안 2(2문항)", 4)

write_header(ws1, ["적합도 지수", "방안 1 (4문항)", "방안 2 (2문항)", "기준"], row=3)
fit_data = [
    ["χ²(df)", "1353.408(206)", "762.764(167)", "-"],
    ["χ²/df", "6.570", "4.567", "≤3 양호, ≤5 수용"],
    ["GFI", ".913", ".943", "≥.90"],
    ["CFI", ".913", ".952", "≥.90"],
    ["TLI", ".903", ".945", "≥.90"],
    ["RMSEA", ".065", ".052", "≤.08"],
    ["RMSEA 90% CI", "[.062, .068]", "[.048, .056]", "-"],
    ["SRMR", ".051", ".039", "≤.08"],
]
last = write_data(ws1, fit_data, start_row=4, num_cols=4, bold_first=True)
add_note(ws1, "주. 방안 2가 모든 적합도 지수에서 우수", last + 2, 4)
add_note(ws1, "    ML 추정, 측정모형 단독 CFA", last + 3, 4)
ws1.column_dimensions["A"].width = 16
for col in "BCD":
    ws1.column_dimensions[col].width = 18


# ============================================================
# Sheet 2: 표준화 요인부하량 (표 11)
# ============================================================
ws2 = wb.create_sheet("2. Factor Loadings (Table 11)")
add_title(ws2, "표준화 요인부하량: 방안 1 vs 방안 2", 4)

write_header(ws2, ["잠재변수", "관측변수", "방안 1 (4문항)", "방안 2 (2문항)"], row=3)
load_data = [
    ["자원봉사 효능감", "효능감1", ".685", ".684"],
    ["", "효능감2", ".619", ".618"],
    ["", "효능감3", ".685", ".685"],
    ["", "효능감4", ".654", ".655"],
    ["", "효능감5", ".662", ".661"],
    ["", "효능감6", ".622", ".622"],
    ["", "효능감7", ".712", ".713"],
    ["", "효능감8", ".660", ".661"],
    ["", "효능감9", ".692", ".692"],
    ["", "효능감10", ".693", ".693"],
    ["자원봉사제도 인식", "제도인식1", ".734", ".734"],
    ["", "제도인식2", ".781", ".781"],
    ["", "제도인식3", ".627", ".626"],
    ["", "제도인식4", ".805", ".805"],
    ["", "제도인식5", ".735", ".735"],
    ["", "제도인식6", ".815", ".816"],
    ["", "제도인식7", ".774", ".774"],
    ["", "제도인식8", ".729", ".729"],
    ["주관적 안녕감", "삶의 만족", ".799", ".910"],
    ["", "행복", ".748", ".686"],
    ["", "걱정(역)", ".375", "(제외)"],
    ["", "우울(역)", ".457", "(제외)"],
]
# 안녕감 행 하이라이트 (18~21)
last = write_data(ws2, load_data, start_row=4, num_cols=4, bold_first=False,
                  hl_rows={18, 19, 20, 21})
add_note(ws2, "주. 모든 요인부하량 p < .001로 통계적으로 유의",
         last + 2, 4)
add_note(ws2, "    방안 1 범위: .38~.82 (걱정 .375 최저) | 방안 2 범위: .62~.91",
         last + 3, 4)
add_note(ws2, "    노란색 음영: 안녕감 측정 변경 부분",
         last + 4, 4)
ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 16


# ============================================================
# Sheet 3: 잠재변수 간 상관 + AVE/CR
# ============================================================
ws3 = wb.create_sheet("3. Correlation & Validity")
add_title(ws3, "잠재변수 간 상관관계 및 집중타당도", 4)

# 상관
ws3.cell(row=3, column=1, value="[A] 잠재변수 간 상관관계").font = header_font
ws3.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
write_header(ws3, ["잠재변수 쌍", "방안 1", "방안 2", "비고"], row=4)
cor_data = [
    ["효능감 ↔ 제도인식", ".450", ".450", "동일"],
    ["효능감 ↔ 안녕감", ".407", ".397", "미세 감소"],
    ["제도인식 ↔ 안녕감", ".269", ".255", "미세 감소"],
]
last = write_data(ws3, cor_data, start_row=5, num_cols=4, bold_first=True)

# AVE/CR
ws3.cell(row=last+2, column=1, value="[B] 집중타당도 (AVE, CR)").font = header_font
ws3.merge_cells(start_row=last+2, start_column=1, end_row=last+2, end_column=4)
write_header(ws3, ["잠재변수", "방안 1 (AVE/CR)", "방안 2 (AVE/CR)", "기준"], row=last+3)
ave_data = [
    ["자원봉사 효능감", ".447 / .890", ".447 / .890", "AVE≥.50, CR≥.70"],
    ["자원봉사제도 인식", ".566 / .912", ".566 / .912", "AVE≥.50, CR≥.70"],
    ["주관적 안녕감", ".387 / .701", ".649 / .784", "AVE≥.50, CR≥.70"],
]
last2 = write_data(ws3, ave_data, start_row=last+4, num_cols=4, bold_first=True,
                   hl_rows={2})
add_note(ws3, "주. 안녕감 AVE: .387(미달) → .649(충족)로 개선 (노란 음영)",
         last2 + 2, 4)
add_note(ws3, "    판별타당도(Fornell-Larcker): 두 방안 모두 모든 쌍 충족",
         last2 + 3, 4)
ws3.column_dimensions["A"].width = 18
for col in "BCD":
    ws3.column_dimensions[col].width = 18


wb.save("results/CFA_option2_measurement.xlsx")
print("저장 완료: results/CFA_option2_measurement.xlsx")
print("총 3개 시트:")
print("  1. Model Fit (Table 9) - 측정모형 적합도 비교")
print("  2. Factor Loadings (Table 11) - 표준화 요인부하량")
print("  3. Correlation & Validity - 잠재변수 상관 + AVE/CR")
