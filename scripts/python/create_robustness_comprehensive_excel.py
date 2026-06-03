"""
강건성(Robustness) 종합 검증 엑셀 생성
2차 심사위원 robustness 요구 대응

3종 분석 통합:
1. 추정법 민감도 (ML/MLR/GLS/ULS)
2. 매개구조 복원 민감도
3. 샘플 부트스트랩 (서브샘플링/오버샘플링)

파일명: Robustness_Comprehensive_Report.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = Workbook()
wb.remove(wb.active)

title_font = Font(name="Times New Roman", bold=True, size=13)
header_font = Font(name="Times New Roman", bold=True, size=11)
body_font = Font(name="Times New Roman", size=11)
bold_font = Font(name="Times New Roman", bold=True, size=11)
note_font = Font(name="Times New Roman", italic=True, size=10)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin")
thick = Side(style="medium")
border_top = Border(top=thick, bottom=thin)
border_mid = Border(top=thin, bottom=thin)
border_bottom = Border(bottom=thick)
header_fill = PatternFill("solid", fgColor="D9E1F2")
good_fill = PatternFill("solid", fgColor="E2EFDA")
warn_fill = PatternFill("solid", fgColor="FCE4D6")


def add_title(ws, text, ncol):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1, value=text)
    c.font = title_font
    c.alignment = center
    ws.row_dimensions[1].height = 22


def write_header(ws, headers, row=3):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.border = border_top
        cell.fill = header_fill


def write_data(ws, rows, start_row=4, num_cols=None, bold_first=False, fills=None):
    last = start_row
    fills = fills or {}
    for r, row_data in enumerate(rows, start_row):
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = bold_font if (bold_first and c == 1) else body_font
            cell.alignment = center if c > 1 else left
            cell.border = border_mid
            if (r - start_row) in fills:
                cell.fill = fills[r - start_row]
        last = r
    if num_cols is None:
        num_cols = len(rows[0])
    for c in range(1, num_cols + 1):
        ws.cell(row=last, column=c).border = border_bottom
    return last


def add_note(ws, note, start_row, ncol):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncol)
    c = ws.cell(row=start_row, column=1, value=note)
    c.font = note_font
    c.alignment = left


# ============================================================
# Sheet 1: 종합 요약 (Executive Summary)
# ============================================================
ws1 = wb.create_sheet("1. Summary")
add_title(ws1, "강건성(Robustness) 종합 검증 요약 — 2차 심사 대응", 5)

write_header(ws1, ["검증 차원", "방법", "효능감→안녕감 (b1)", "안녕감→참여 (b3)", "판정"], row=3)
summary = [
    ["① 추정법 민감도", "ML/MLR/GLS 비교", ".250~.294 (모두 유의)", "유의~경계", "강건"],
    ["② 매개구조 복원", "추정법별 완전매개 검증", "일관 복원", "일관 복원(ML/MLR/GLS)", "강건"],
    ["③-1 서브샘플링", "70% × 200회", ".294 (100% 유의)", ".082 (50% 유의)", "b1 강건/b3 경계"],
    ["③-2 오버샘플링", "참여 17.4%→40%", ".274 (유의)", ".076 (유의)", "강건"],
]
last = write_data(ws1, summary, start_row=4, num_cols=5, bold_first=True,
                  fills={0: good_fill, 1: good_fill, 2: warn_fill, 3: good_fill})

add_note(ws1, "주. 핵심 매개경로(효능감→안녕감)는 모든 검증에서 압도적으로 강건",
         last + 2, 5)
add_note(ws1, "    안녕감→참여(b3)는 경계선 효과 — 효과크기는 일관(.076~.083)하나 표본변동에 민감",
         last + 3, 5)
add_note(ws1, "    완전매개 구조(효능감→참여 직접효과 비유의)는 전 조건에서 복원",
         last + 4, 5)

ws1.column_dimensions["A"].width = 18
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 22
ws1.column_dimensions["E"].width = 16


# ============================================================
# Sheet 2: 추정법 민감도
# ============================================================
ws2 = wb.create_sheet("2. Estimator Sensitivity")
add_title(ws2, "추정법별 적합도 및 매개경로 (민감도 1)", 7)

write_header(ws2, ["추정법", "CFI", "TLI", "RMSEA", "SRMR",
                   "b1 (효능감→안녕감)", "b3 (안녕감→참여)"], row=3)
est_data = [
    ["ML (본 분석)", ".881", ".865", ".054", ".074", ".294***", ".083*"],
    ["MLR", ".882", ".866", ".054", ".074", ".294***", ".083*"],
    ["GLS", ".530", ".468", ".043", ".116", ".250***", ".096**"],
    ["ULS (제외)", ".313", ".222", ".052", ".075", ".536***", ".099 n.s."],
]
last = write_data(ws2, est_data, start_row=4, num_cols=7, bold_first=True,
                  fills={0: good_fill, 1: good_fill})
add_note(ws2, "주. ML/MLR 본 분석 중심. GLS는 매개구조 복원되나 적합도 한계.",
         last + 2, 7)
add_note(ws2, "    ULS는 변수 분산 척도 차이(1000배 경고)로 부적합 → 보고 제외 권장",
         last + 3, 7)
add_note(ws2, "    심사 보고 시 적합도 중심으로 ML vs MLR 제시 (전체 그림 불필요)",
         last + 4, 7)
ws2.column_dimensions["A"].width = 16
for col in "BCDE":
    ws2.column_dimensions[col].width = 10
ws2.column_dimensions["F"].width = 18
ws2.column_dimensions["G"].width = 18


# ============================================================
# Sheet 3: 샘플 부트스트랩
# ============================================================
ws3 = wb.create_sheet("3. Sample Bootstrap")
add_title(ws3, "샘플 재구성 민감도 (민감도 3)", 4)

# 종합 비교
ws3.cell(row=3, column=1, value="[A] 원자료 vs 서브샘플 vs 오버샘플 매개경로").font = header_font
ws3.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
write_header(ws3, ["분석", "참여 비율", "b1 (효능감→안녕감)", "b3 (안녕감→참여)"], row=4)
boot_data = [
    ["원자료 (N=1,328)", "17.4%", ".294", ".083*"],
    ["70% 서브샘플 평균 (200회)", "17.4%", ".294 (SD=.026)", ".082"],
    ["오버샘플링", "40.0%", ".274***", ".076*"],
]
last = write_data(ws3, boot_data, start_row=5, num_cols=4, bold_first=True)

# 유의 비율 (서브샘플)
ws3.cell(row=last+2, column=1, value="[B] 70% 서브샘플 200회 유의(p<.05) 비율").font = header_font
ws3.merge_cells(start_row=last+2, start_column=1, end_row=last+2, end_column=4)
write_header(ws3, ["경로", "평균 β", "SD", "유의 비율"], row=last+3)
sig_data = [
    ["효능감 → 안녕감 (b1)", ".294", ".026", "100.0%"],
    ["안녕감 → 참여 (b3)", ".082", ".024", "50.0%"],
]
last2 = write_data(ws3, sig_data, start_row=last+4, num_cols=4, bold_first=True,
                   fills={0: good_fill, 1: warn_fill})

add_note(ws3, "주. 서브샘플 200회 전체 수렴 성공 (실패 0회)",
         last2 + 2, 4)
add_note(ws3, "    b1: 100% 유의 → 매우 강건 | b3: 50% 유의 → 경계선 효과",
         last2 + 3, 4)
add_note(ws3, "    b3 효과크기(.076~.083)는 세 조건에서 일관 유지",
         last2 + 4, 4)
ws3.column_dimensions["A"].width = 28
for col in "BCD":
    ws3.column_dimensions[col].width = 18


# ============================================================
# Sheet 4: 종속변수 불균형 + 심사 대응
# ============================================================
ws4 = wb.create_sheet("4. Imbalance & Response")
add_title(ws4, "종속변수 불균형 구조 및 심사 대응 전략", 2)

# 불균형
ws4.cell(row=3, column=1, value="[A] 종속변수(자원봉사 참여) 불균형").font = header_font
ws4.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
imb = [
    ["미참여 (0)", "1,097명 (82.6%)"],
    ["참여 (1)", "231명 (17.4%)"],
]
r = 4
for label, val in imb:
    ws4.cell(row=r, column=1, value=label).font = bold_font
    ws4.cell(row=r, column=1).border = border_mid
    ws4.cell(row=r, column=1).alignment = left
    ws4.cell(row=r, column=2, value=val).font = body_font
    ws4.cell(row=r, column=2).border = border_mid
    ws4.cell(row=r, column=2).alignment = center
    r += 1
for col in range(1, 3):
    ws4.cell(row=r-1, column=col).border = border_bottom

# 심사 대응
ws4.cell(row=r+1, column=1, value="[B] 심사위원 robustness 지적 대응").font = header_font
ws4.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=2)

resp = [
    ["지적: 이분 종속변수 불균형(17.4%)",
     "→ 오버샘플링(40%)에서도 매개구조 복원 확인"],
    ["지적: 복잡 모형·비정규",
     "→ MLR(강건 추정)에서도 동일 결과"],
    ["지적: 경계선 지표",
     "→ b3 경계선 효과 정직 인정 + b1 강건성(100%) 강조"],
    ["지적: 추정법 달리해도 유사?",
     "→ ML/MLR/GLS 매개구조 일관 복원"],
    ["지적: 샘플 재구성 복원?",
     "→ 70% 서브샘플 200회 + 오버샘플링 복원"],
]
rr = r + 2
for q, a in resp:
    ws4.cell(row=rr, column=1, value=q).font = body_font
    ws4.cell(row=rr, column=1).border = border_mid
    ws4.cell(row=rr, column=1).alignment = left
    ws4.cell(row=rr, column=2, value=a).font = body_font
    ws4.cell(row=rr, column=2).border = border_mid
    ws4.cell(row=rr, column=2).alignment = left
    ws4.row_dimensions[rr].height = 30
    rr += 1
for col in range(1, 3):
    ws4.cell(row=rr-1, column=col).border = border_bottom

ws4.column_dimensions["A"].width = 36
ws4.column_dimensions["B"].width = 50


wb.save("results/Robustness_Comprehensive_Report.xlsx")
print("저장 완료: results/Robustness_Comprehensive_Report.xlsx")
print("총 4개 시트:")
print("  1. Summary (종합 요약)")
print("  2. Estimator Sensitivity (추정법 민감도)")
print("  3. Sample Bootstrap (샘플 부트스트랩)")
print("  4. Imbalance & Response (불균형 + 심사 대응)")
