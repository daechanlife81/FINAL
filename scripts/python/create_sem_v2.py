from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

wb = Workbook()

header_font = Font(name="Times New Roman", bold=True, size=11)
body_font = Font(name="Times New Roman", size=11)
bold_font = Font(name="Times New Roman", bold=True, size=11)
title_font = Font(name="Times New Roman", bold=True, size=13)
note_font = Font(name="Times New Roman", italic=True, size=10)
center = Alignment(horizontal="center", vertical="center")
center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center")
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin")
thick = Side(style="medium")
border_top = Border(top=thick, bottom=thin)
border_mid = Border(top=thin, bottom=thin)
border_bottom = Border(bottom=thick)

# ========== Sheet 1: 구조모형 적합도 ==========
ws1 = wb.active
ws1.title = "구조모형 적합도"

ws1.merge_cells("A1:C1")
ws1["A1"] = "구조모형 적합도 검증 결과"
ws1["A1"].font = title_font
ws1["A1"].alignment = center

for i, h in enumerate(["적합도 지수", "값", "기준"], 1):
    cell = ws1.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

fit_data = [
    ("chi2(df)", "2165.325(447), p<.001", "-"),
    ("chi2/df", "4.844", "-"),
    ("RMSEA [90% CI]", ".054 [.052, .056]", "<=.06"),
    ("SRMR", ".074", "<=.09"),
    ("CFI", ".881", "close to .95"),
    ("TLI", ".865", "close to .95"),
]

for r, (a, b, c) in enumerate(fit_data, 4):
    for col, v in enumerate([a, b, c], 1):
        cell = ws1.cell(row=r, column=col, value=v)
        cell.font = body_font
        cell.alignment = center
        cell.border = border_mid

for c in range(1, 4):
    ws1.cell(row=9, column=c).border = border_bottom

ws1.merge_cells("A10:C10")
ws1["A10"] = "주. 적합도 기준: Hu & Bentler(1999). 2-지수 전략(RMSEA+SRMR) 충족."
ws1["A10"].font = note_font
ws1.merge_cells("A11:C11")
ws1["A11"] = "    CFI/TLI는 변수 수 증가 시 악화 경향 있음(Kenny & McCoach, 2003)."
ws1["A11"].font = note_font

ws1.column_dimensions["A"].width = 20
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 20

# ========== Sheet 2: 구조경로 계수 ==========
ws2 = wb.create_sheet("구조경로 계수")

ws2.merge_cells("A1:H1")
ws2["A1"] = "구조모형 경로계수 추정 결과"
ws2["A1"].font = title_font
ws2["A1"].alignment = center

headers2 = ["구분", "경로", "B", "S.E.", "beta", "p", "", ""]
headers2_display = ["구분", "경로", "B", "S.E.", "\u03b2", "p", ""]
for i, h in enumerate(headers2_display, 1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

path_data = [
    ("연구문제1\n독립\u2192효능감", "주관적 건강 \u2192 효능감", ".094", ".019", ".144", "<.001", "***"),
    ("", "교육수준 \u2192 효능감", ".045", ".021", ".069", ".033", "*"),
    ("", "일반적 신뢰 \u2192 효능감", ".035", ".022", ".042", ".114", ""),
    ("", "네트워크 \u2192 효능감", ".113", ".038", ".078", ".003", "**"),
    ("", "사회단체활동 \u2192 효능감", ".003", ".030", ".003", ".926", ""),
    ("", "제도인식 \u2192 효능감", ".232", ".021", ".332", "<.001", "***"),
    ("", "종교활동 \u2192 효능감", ".087", ".031", ".079", ".006", "**"),
    ("", "봉사교육경험 \u2192 효능감", ".177", ".032", ".156", "<.001", "***"),
    ("연구문제2\n독립\u2192안녕감", "주관적 건강 \u2192 안녕감", ".301", ".043", ".230", "<.001", "***"),
    ("", "교육수준 \u2192 안녕감", ".047", ".046", ".036", ".303", ""),
    ("", "일반적 신뢰 \u2192 안녕감", ".449", ".049", ".269", "<.001", "***"),
    ("", "네트워크 \u2192 안녕감", ".220", ".083", ".075", ".009", "**"),
    ("", "사회단체활동 \u2192 안녕감", ".077", ".065", ".037", ".235", ""),
    ("", "종교활동 \u2192 안녕감", "-.049", ".068", "-.022", ".473", ""),
    ("연구문제3\n독립\u2192참여", "주관적 건강 \u2192 참여", "-.032", ".015", "-.062", ".036", "*"),
    ("", "교육수준 \u2192 참여", "-.014", ".016", "-.028", ".369", ""),
    ("", "일반적 신뢰 \u2192 참여", ".021", ".018", ".032", ".236", ""),
    ("", "네트워크 \u2192 참여", "-.000", ".029", "-.000", ".996", ""),
    ("", "사회단체활동 \u2192 참여", ".080", ".023", ".097", "<.001", "***"),
    ("", "제도인식 \u2192 참여", ".127", ".016", ".231", "<.001", "***"),
    ("", "종교활동 \u2192 참여", ".078", ".024", ".091", ".001", "***"),
    ("", "봉사교육경험 \u2192 참여", ".049", ".024", ".055", ".045", "*"),
    ("연구문제4\n매개경로", "효능감 \u2192 안녕감", ".588", ".068", ".294", "<.001", "***"),
    ("", "효능감 \u2192 참여", ".004", ".026", ".005", ".876", ""),
    ("", "안녕감 \u2192 참여", ".033", ".014", ".083", ".017", "*"),
    ("통제변수\n\u2192효능감", "성별 \u2192 효능감", ".026", ".025", ".028", ".293", ""),
    ("", "연령 \u2192 효능감", ".002", ".001", ".055", ".170", ""),
    ("", "혼인상태 \u2192 효능감", "-.001", ".027", "-.002", ".958", ""),
    ("", "과거봉사경험 \u2192 효능감", ".182", ".028", ".191", "<.001", "***"),
    ("통제변수\n\u2192안녕감", "성별 \u2192 안녕감", ".031", ".054", ".016", ".565", ""),
    ("", "연령 \u2192 안녕감", ".003", ".003", ".052", ".233", ""),
    ("", "혼인상태 \u2192 안녕감", "-.001", ".058", "-.001", ".987", ""),
    ("", "과거봉사경험 \u2192 안녕감", "-.024", ".059", "-.013", ".678", ""),
    ("통제변수\n\u2192참여", "성별 \u2192 참여", "-.017", ".019", "-.023", ".371", ""),
    ("", "연령 \u2192 참여", "-.000", ".001", "-.003", ".931", ""),
    ("", "혼인상태 \u2192 참여", ".023", ".020", ".038", ".247", ""),
    ("", "과거봉사경험 \u2192 참여", ".181", ".021", ".242", "<.001", "***"),
]

r = 4
prev_group = None
group_start = None

for grp, path, b, se, beta, p, sig in path_data:
    if grp and grp != prev_group:
        if prev_group and group_start and group_start < r:
            ws2.merge_cells(start_row=group_start, start_column=1, end_row=r-1, end_column=1)
            ws2.cell(row=group_start, column=1).alignment = center_wrap
        group_start = r
        prev_group = grp

    vals = [grp if grp else "", path, b, se, beta, p, sig]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.font = body_font
        cell.alignment = center if c >= 3 else left_wrap
        cell.border = border_mid

    r += 1

# 마지막 그룹 병합
if group_start and group_start < r:
    ws2.merge_cells(start_row=group_start, start_column=1, end_row=r-1, end_column=1)
    ws2.cell(row=group_start, column=1).alignment = center_wrap

last_row = r - 1
for c in range(1, 8):
    ws2.cell(row=last_row, column=c).border = border_bottom

# R-squared
r2_row = last_row + 1
ws2.merge_cells(f"A{r2_row}:G{r2_row}")
ws2[f"A{r2_row}"] = "R\u00b2: \ud6a8\ub2a5\uac10=.265, \uc548\ub155\uac10=.276, \ucc38\uc5ec=.202"
ws2[f"A{r2_row}"].font = bold_font

note_row = r2_row + 1
ws2.merge_cells(f"A{note_row}:G{note_row}")
ws2[f"A{note_row}"] = "\uc8fc. *p<.05, **p<.01, ***p<.001."
ws2[f"A{note_row}"].font = note_font

ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 28
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 10
ws2.column_dimensions["E"].width = 10
ws2.column_dimensions["F"].width = 10
ws2.column_dimensions["G"].width = 6

# ========== Sheet 3: 매개효과 ==========
ws3 = wb.create_sheet("매개효과")

ws3.merge_cells("A1:G1")
ws3["A1"] = "매개효과 검증 결과 (Bootstrap 5,000회, 95% CI)"
ws3["A1"].font = title_font
ws3["A1"].alignment = center

headers3 = ["매개유형", "경로", "B", "\u03b2", "95% CI 하한", "95% CI 상한", ""]
for i, h in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

med_data = [
    ("단순매개\n독립\u2192효능감\u2192참여", "주관적건강\u2192효능감\u2192참여", ".0004", ".001", "-.005", ".005", ""),
    ("", "교육수준\u2192효능감\u2192참여", ".0002", ".000", "-.002", ".003", ""),
    ("", "일반적신뢰\u2192효능감\u2192참여", ".0001", ".000", "-.002", ".003", ""),
    ("", "네트워크\u2192효능감\u2192참여", ".0005", ".000", "-.005", ".007", ""),
    ("", "사회단체활동\u2192효능감\u2192참여", ".0000", ".000", "-.002", ".002", ""),
    ("", "제도인식\u2192효능감\u2192참여", ".0009", ".002", "-.011", ".013", ""),
    ("", "종교활동\u2192효능감\u2192참여", ".0004", ".000", "-.004", ".005", ""),
    ("", "봉사교육경험\u2192효능감\u2192참여", ".0007", ".001", "-.008", ".010", ""),
    ("단순매개\n독립\u2192안녕감\u2192참여", "주관적건강\u2192안녕감\u2192참여", ".0098", ".019", ".001", ".020", "*"),
    ("", "교육수준\u2192안녕감\u2192참여", ".0015", ".003", "-.002", ".006", ""),
    ("", "일반적신뢰\u2192안녕감\u2192참여", ".0147", ".022", ".002", ".028", "*"),
    ("", "네트워크\u2192안녕감\u2192참여", ".0072", ".006", ".000", ".017", "*"),
    ("", "사회단체활동\u2192안녕감\u2192참여", ".0025", ".003", "-.002", ".009", ""),
    ("", "종교활동\u2192안녕감\u2192참여", "-.0016", "-.002", "-.008", ".003", ""),
    ("순차매개\n독립\u2192효능감\n\u2192안녕감\u2192참여", "주관적건강\u2192효능감\u2192안녕감\u2192참여", ".0018", ".004", ".000", ".004", "*"),
    ("", "교육수준\u2192효능감\u2192안녕감\u2192참여", ".0009", ".002", "-.000", ".002", ""),
    ("", "일반적신뢰\u2192효능감\u2192안녕감\u2192참여", ".0007", ".001", "-.000", ".002", ""),
    ("", "네트워크\u2192효능감\u2192안녕감\u2192참여", ".0022", ".002", ".000", ".005", "*"),
    ("", "사회단체활동\u2192효능감\u2192안녕감\u2192참여", ".0001", ".000", "-.001", ".001", ""),
    ("", "제도인식\u2192효능감\u2192안녕감\u2192참여", ".0045", ".008", ".001", ".009", "*"),
    ("", "종교활동\u2192효능감\u2192안녕감\u2192참여", ".0017", ".002", ".000", ".004", "*"),
    ("", "봉사교육경험\u2192효능감\u2192안녕감\u2192참여", ".0034", ".004", ".000", ".007", "*"),
]

r = 4
prev_group = None
group_start = None

for mtype, path, b, beta, ci_lo, ci_hi, sig in med_data:
    if mtype and mtype != prev_group:
        if prev_group and group_start and group_start < r:
            ws3.merge_cells(start_row=group_start, start_column=1, end_row=r-1, end_column=1)
            ws3.cell(row=group_start, column=1).alignment = center_wrap
        group_start = r
        prev_group = mtype

    vals = [mtype if mtype else "", path, b, beta, ci_lo, ci_hi, sig]
    for c, v in enumerate(vals, 1):
        cell = ws3.cell(row=r, column=c, value=v)
        cell.font = body_font
        cell.alignment = center_wrap if c == 1 else center
        cell.border = border_mid

    r += 1

if group_start and group_start < r:
    ws3.merge_cells(start_row=group_start, start_column=1, end_row=r-1, end_column=1)
    ws3.cell(row=group_start, column=1).alignment = center_wrap

last_row3 = r - 1
for c in range(1, 8):
    ws3.cell(row=last_row3, column=c).border = border_bottom

note_row3 = last_row3 + 1
ws3.merge_cells(f"A{note_row3}:G{note_row3}")
ws3[f"A{note_row3}"] = "주. *95% 신뢰구간에 0이 포함되지 않으면 유의한 매개효과."
ws3[f"A{note_row3}"].font = note_font

ws3.column_dimensions["A"].width = 18
ws3.column_dimensions["B"].width = 38
ws3.column_dimensions["C"].width = 10
ws3.column_dimensions["D"].width = 10
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 14
ws3.column_dimensions["G"].width = 6

# ========== Sheet 4: 가설검증 종합 ==========
ws4 = wb.create_sheet("가설검증 종합")

ws4.merge_cells("A1:F1")
ws4["A1"] = "연구가설 검증 결과 종합"
ws4["A1"].font = title_font
ws4["A1"].alignment = center

headers4 = ["가설", "내용", "판단"]
for i, h in enumerate(headers4, 1):
    cell = ws4.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border_top

hyp_data = [
    ("H1", "인간자본(건강, 교육) \u2192 효능감", "채택"),
    ("H2", "사회자본(신뢰, 네트워크, 사회단체) \u2192 효능감", "부분채택"),
    ("H3", "문화자본(제도인식, 종교, 봉사교육) \u2192 효능감", "채택"),
    ("H4", "인간자본(건강, 교육) \u2192 안녕감", "부분채택"),
    ("H5", "사회자본(신뢰, 네트워크, 사회단체) \u2192 안녕감", "부분채택"),
    ("H6", "종교활동 \u2192 안녕감", "기각"),
    ("H7", "인간자본(건강, 교육) \u2192 참여", "기각"),
    ("H8", "사회자본(신뢰, 네트워크, 사회단체) \u2192 참여", "부분채택"),
    ("H9", "문화자본(제도인식, 종교, 봉사교육) \u2192 참여", "부분채택"),
    ("H10", "효능감 \u2192 안녕감", "채택"),
    ("H11", "효능감 \u2192 참여", "기각"),
    ("H12", "안녕감 \u2192 참여", "채택"),
    ("H13", "독립 \u2192 효능감 \u2192 참여 (단순매개)", "기각"),
    ("H14", "독립 \u2192 안녕감 \u2192 참여 (단순매개)", "부분채택"),
    ("H15", "독립 \u2192 효능감 \u2192 안녕감 \u2192 참여 (순차매개)", "부분채택"),
]

for r, (h, content, result) in enumerate(hyp_data, 4):
    for c, v in enumerate([h, content, result], 1):
        cell = ws4.cell(row=r, column=c, value=v)
        cell.font = bold_font if c == 3 else body_font
        cell.alignment = center
        cell.border = border_mid

last_row4 = 4 + len(hyp_data) - 1
for c in range(1, 4):
    ws4.cell(row=last_row4, column=c).border = border_bottom

summary_row = last_row4 + 1
ws4.merge_cells(f"A{summary_row}:C{summary_row}")
ws4[f"A{summary_row}"] = "종합: 채택 4개, 부분채택 7개, 기각 4개 (총 15개)"
ws4[f"A{summary_row}"].font = bold_font
ws4[f"A{summary_row}"].alignment = center

ws4.column_dimensions["A"].width = 8
ws4.column_dimensions["B"].width = 50
ws4.column_dimensions["C"].width = 14

wb.save("SEM_results_v2.xlsx")
print("saved: SEM_results_v2.xlsx")
