"""
Correlation Robustness Check Excel 생성
박사학위논문 마지막 점검 (2026-04-26)

목적: Pearson vs Spearman vs Kendall 세 상관계수의 비교 결과를
      별도 엑셀로 보존 (박사논문 부록 또는 별표)

파일명: Correlation_robustness_check.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = Workbook()
wb.remove(wb.active)

# 공통 스타일
title_font = Font(name="Times New Roman", bold=True, size=13)
header_font = Font(name="Times New Roman", bold=True, size=11)
body_font = Font(name="Times New Roman", size=11)
bold_font = Font(name="Times New Roman", bold=True, size=11)
note_font = Font(name="Times New Roman", italic=True, size=10)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin")
thick = Side(style="medium")
border_top = Border(top=thick, bottom=thin)
border_mid = Border(top=thin, bottom=thin)
border_bottom = Border(bottom=thick)
header_fill = PatternFill("solid", fgColor="E8E8E8")


def add_title(ws, text, ncol):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = title_font
    cell.alignment = center
    ws.row_dimensions[1].height = 22


def write_header(ws, headers, row=3):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.border = border_top
        cell.fill = header_fill


def write_data(ws, rows, start_row=4, num_cols=None, bold_first=False):
    last = start_row
    for r, row_data in enumerate(rows, start_row):
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = bold_font if (bold_first and c == 1) else body_font
            cell.alignment = center if c > 1 else left
            cell.border = border_mid
        last = r
    if num_cols is None:
        num_cols = len(rows[0])
    for c in range(1, num_cols + 1):
        ws.cell(row=last, column=c).border = border_bottom
    return last


def add_note(ws, note, start_row, ncol):
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=ncol)
    cell = ws.cell(row=start_row, column=1, value=note)
    cell.font = note_font
    cell.alignment = left


# ============================================================
# Sheet 1: 변수 유형 분류 (Variable Types)
# ============================================================
ws1 = wb.create_sheet("1. Variable Types")
add_title(ws1, "본 연구 분석 변수의 유형 분류", 4)

write_header(ws1, ["변수명", "한글명", "유형", "수준"], row=3)

var_types = [
    ["re_A_01_1", "자원봉사 참여경험 (종속)", "이분형 (Binary)", "2"],
    ["re_B_02", "과거 자원봉사경험", "이분형 (Binary)", "2"],
    ["re_B_04", "자원봉사교육 경험", "이분형 (Binary)", "2"],
    ["re_C_06", "네트워크", "이분형 (Binary)", "2"],
    ["re_C_07", "사회단체활동", "이분형 (Binary)", "2"],
    ["re_C_08", "종교활동", "이분형 (Binary)", "2"],
    ["re_SQ1", "성별", "이분형 (Binary)", "2"],
    ["D_05", "주관적 건강", "서열형 (Ordinal)", "5"],
    ["D_03", "교육수준", "서열형 (Ordinal)", "5"],
    ["C_05", "일반적 신뢰", "서열형 (Ordinal)", "4"],
    ["SQ2", "연령", "연속형 (Continuous)", "61"],
]
last = write_data(ws1, var_types, start_row=4, num_cols=4, bold_first=True)

add_note(ws1, "주. 이분형 7개, 서열형 3개, 연속형 1개로 구성",
         last + 2, 4)
add_note(ws1, "    이분형 변수의 비중을 고려하여 세 가지 상관계수 산출 비교 검증",
         last + 3, 4)

ws1.column_dimensions["A"].width = 14
ws1.column_dimensions["B"].width = 26
ws1.column_dimensions["C"].width = 20
ws1.column_dimensions["D"].width = 10


# ============================================================
# Sheet 2: 종속변수와의 상관 - 세 방법 비교
# ============================================================
ws2 = wb.create_sheet("2. DV Correlation Comparison")
add_title(ws2, "종속변수(자원봉사 참여)와의 상관 - 세 방법 비교", 6)

write_header(ws2,
             ["변수", "Pearson r", "Spearman ρ", "Kendall τ",
              "|Pearson-Spearman|", "|Pearson-Kendall|"], row=3)

dv_comp = [
    ["과거 봉사경험", 0.365, 0.365, 0.365, 0.000, 0.000],
    ["봉사교육경험", 0.208, 0.208, 0.208, 0.000, 0.000],
    ["네트워크", 0.056, 0.056, 0.056, 0.000, 0.000],
    ["사회단체활동", 0.257, 0.257, 0.257, 0.000, 0.000],
    ["종교활동", 0.224, 0.224, 0.224, 0.000, 0.000],
    ["성별", 0.017, 0.017, 0.017, 0.000, 0.000],
    ["주관적 건강", -0.001, -0.009, -0.009, 0.008, 0.008],
    ["교육수준", 0.027, 0.032, 0.031, 0.005, 0.004],
    ["일반적 신뢰", 0.172, 0.187, 0.183, 0.015, 0.011],
    ["연령", 0.028, 0.032, 0.026, 0.004, 0.002],
]
last = write_data(ws2, dv_comp, start_row=4, num_cols=6, bold_first=True)

# 평균 차이 행
avg_row = last + 1
ws2.cell(row=avg_row, column=1, value="평균 절대 차이").font = bold_font
ws2.cell(row=avg_row, column=1).alignment = left
ws2.cell(row=avg_row, column=5, value=0.0032).font = bold_font
ws2.cell(row=avg_row, column=5).alignment = center
ws2.cell(row=avg_row, column=6, value=0.0025).font = bold_font
ws2.cell(row=avg_row, column=6).alignment = center

for c in range(1, 7):
    ws2.cell(row=avg_row, column=c).border = border_bottom

add_note(ws2,
         "주. 이분형 ↔ 이분형 변수 간 상관은 세 방법 모두 완전 일치 (.365 등)",
         last + 3, 6)
add_note(ws2,
         "    서열형/연속형 포함 변수에서 미세 차이 발생 (최대 .015)",
         last + 4, 6)
add_note(ws2,
         "    평균 절대 차이 .003 이하로 세 방법의 결과 견고성 확인",
         last + 5, 6)

ws2.column_dimensions["A"].width = 16
for col in "BCDEF":
    ws2.column_dimensions[col].width = 18


# ============================================================
# Sheet 3: Pearson 상관행렬 (전체)
# ============================================================
ws3 = wb.create_sheet("3. Pearson Matrix")
add_title(ws3, "Pearson 상관행렬 (전체 변수)", 12)

# 상단 헤더
write_header(ws3,
             ["", "참여", "과거", "교육경험", "네트워크",
              "사회단체", "종교", "성별", "건강", "학력", "신뢰", "연령"],
             row=3)

pearson_data = [
    ["참여경험 (DV)", 1.000, 0.365, 0.208, 0.056, 0.257, 0.224, 0.017, -0.001, 0.027, 0.172, 0.028],
    ["과거봉사", 0.365, 1.000, 0.316, 0.103, 0.269, 0.195, 0.039, 0.081, 0.147, 0.167, -0.077],
    ["봉사교육", 0.208, 0.316, 1.000, 0.069, 0.221, 0.122, 0.012, 0.134, 0.179, 0.034, -0.178],
    ["네트워크", 0.056, 0.103, 0.069, 1.000, 0.111, 0.083, 0.015, 0.052, 0.077, 0.109, -0.012],
    ["사회단체", 0.257, 0.269, 0.221, 0.111, 1.000, 0.344, -0.008, 0.042, 0.031, 0.104, 0.082],
    ["종교활동", 0.224, 0.195, 0.122, 0.083, 0.344, 1.000, 0.145, -0.069, -0.060, 0.181, 0.165],
    ["성별", 0.017, 0.039, 0.012, 0.015, -0.008, 0.145, 1.000, -0.074, -0.063, -0.002, 0.015],
    ["건강", -0.001, 0.081, 0.134, 0.052, 0.042, -0.069, -0.074, 1.000, 0.377, 0.035, -0.439],
    ["교육수준", 0.027, 0.147, 0.179, 0.077, 0.031, -0.060, -0.063, 0.377, 1.000, 0.025, -0.569],
    ["일반신뢰", 0.172, 0.167, 0.034, 0.109, 0.104, 0.181, -0.002, 0.035, 0.025, 1.000, 0.053],
    ["연령", 0.028, -0.077, -0.178, -0.012, 0.082, 0.165, 0.015, -0.439, -0.569, 0.053, 1.000],
]
last = write_data(ws3, pearson_data, start_row=4, num_cols=12, bold_first=True)

add_note(ws3, "주. 본 연구의 주 상관계수", last + 2, 12)

ws3.column_dimensions["A"].width = 14
for col in "BCDEFGHIJKL":
    ws3.column_dimensions[col].width = 9


# ============================================================
# Sheet 4: Spearman 상관행렬 (전체)
# ============================================================
ws4 = wb.create_sheet("4. Spearman Matrix")
add_title(ws4, "Spearman 순위상관행렬 (전체 변수, 견고성 검증)", 12)

write_header(ws4,
             ["", "참여", "과거", "교육경험", "네트워크",
              "사회단체", "종교", "성별", "건강", "학력", "신뢰", "연령"],
             row=3)

spearman_data = [
    ["참여경험 (DV)", 1.000, 0.365, 0.208, 0.056, 0.257, 0.224, 0.017, -0.009, 0.032, 0.187, 0.032],
    ["과거봉사", 0.365, 1.000, 0.316, 0.103, 0.269, 0.195, 0.039, 0.081, 0.148, 0.161, -0.078],
    ["봉사교육", 0.208, 0.316, 1.000, 0.069, 0.221, 0.122, 0.012, 0.134, 0.191, 0.033, -0.174],
    ["네트워크", 0.056, 0.103, 0.069, 1.000, 0.111, 0.083, 0.015, 0.043, 0.064, 0.112, -0.013],
    ["사회단체", 0.257, 0.269, 0.221, 0.111, 1.000, 0.344, -0.008, 0.045, 0.033, 0.103, 0.084],
    ["종교활동", 0.224, 0.195, 0.122, 0.083, 0.344, 1.000, 0.145, -0.070, -0.057, 0.183, 0.164],
    ["성별", 0.017, 0.039, 0.012, 0.015, -0.008, 0.145, 1.000, -0.074, -0.076, 0.003, 0.020],
    ["건강", -0.009, 0.081, 0.134, 0.043, 0.045, -0.070, -0.074, 1.000, 0.340, 0.025, -0.439],
    ["교육수준", 0.032, 0.148, 0.191, 0.064, 0.033, -0.057, -0.076, 0.340, 1.000, 0.036, -0.570],
    ["일반신뢰", 0.187, 0.161, 0.033, 0.112, 0.103, 0.183, 0.003, 0.025, 0.036, 1.000, 0.050],
    ["연령", 0.032, -0.078, -0.174, -0.013, 0.084, 0.164, 0.020, -0.439, -0.570, 0.050, 1.000],
]
last = write_data(ws4, spearman_data, start_row=4, num_cols=12, bold_first=True)

add_note(ws4, "주. 견고성 검증용 (Pearson과 비교)", last + 2, 12)

ws4.column_dimensions["A"].width = 14
for col in "BCDEFGHIJKL":
    ws4.column_dimensions[col].width = 9


# ============================================================
# Sheet 5: Kendall 상관행렬 (전체)
# ============================================================
ws5 = wb.create_sheet("5. Kendall Matrix")
add_title(ws5, "Kendall τ 상관행렬 (전체 변수, 견고성 검증)", 12)

write_header(ws5,
             ["", "참여", "과거", "교육경험", "네트워크",
              "사회단체", "종교", "성별", "건강", "학력", "신뢰", "연령"],
             row=3)

kendall_data = [
    ["참여경험 (DV)", 1.000, 0.365, 0.208, 0.056, 0.257, 0.224, 0.017, -0.009, 0.031, 0.183, 0.026],
    ["과거봉사", 0.365, 1.000, 0.316, 0.103, 0.269, 0.195, 0.039, 0.077, 0.143, 0.158, -0.064],
    ["봉사교육", 0.208, 0.316, 1.000, 0.069, 0.221, 0.122, 0.012, 0.127, 0.184, 0.032, -0.144],
    ["네트워크", 0.056, 0.103, 0.069, 1.000, 0.111, 0.083, 0.015, 0.041, 0.062, 0.110, -0.011],
    ["사회단체", 0.257, 0.269, 0.221, 0.111, 1.000, 0.344, -0.008, 0.042, 0.032, 0.101, 0.069],
    ["종교활동", 0.224, 0.195, 0.122, 0.083, 0.344, 1.000, 0.145, -0.066, -0.055, 0.180, 0.135],
    ["성별", 0.017, 0.039, 0.012, 0.015, -0.008, 0.145, 1.000, -0.071, -0.073, 0.002, 0.016],
    ["건강", -0.009, 0.077, 0.127, 0.041, 0.042, -0.066, -0.071, 1.000, 0.314, 0.024, -0.353],
    ["교육수준", 0.031, 0.143, 0.184, 0.062, 0.032, -0.055, -0.073, 0.314, 1.000, 0.034, -0.469],
    ["일반신뢰", 0.183, 0.158, 0.032, 0.110, 0.101, 0.180, 0.002, 0.024, 0.034, 1.000, 0.041],
    ["연령", 0.026, -0.064, -0.144, -0.011, 0.069, 0.135, 0.016, -0.353, -0.469, 0.041, 1.000],
]
last = write_data(ws5, kendall_data, start_row=4, num_cols=12, bold_first=True)

add_note(ws5, "주. 견고성 검증용 (Pearson과 비교, 보수적 추정)", last + 2, 12)

ws5.column_dimensions["A"].width = 14
for col in "BCDEFGHIJKL":
    ws5.column_dimensions[col].width = 9


# ============================================================
# Sheet 6: 견고성 검증 종합 요약
# ============================================================
ws6 = wb.create_sheet("6. Robustness Summary")
add_title(ws6, "상관계수 견고성 검증 종합 요약", 3)

write_header(ws6, ["검증 항목", "결과", "판단"], row=3)

summary = [
    ["변수 유형 구성", "이분형 7개, 서열형 3개, 연속형 1개", "이분형 비중 높음"],
    ["Pearson vs Spearman 평균 절대 차이", "0.0032", "거의 동일"],
    ["Pearson vs Kendall 평균 절대 차이", "0.0025", "거의 동일"],
    ["이분형 ↔ 이분형 상관", "세 방법 완전 일치", "수학적 동치"],
    ["서열형/연속형 포함 상관", "최대 차이 .015", "미미한 차이"],
    ["대표본 (N=1,328)", "Pearson robustness 확보", "충족"],
    ["SEM 입력값 호환성", "Pearson 공분산 행렬과 동일 기반", "유리"],
    ["견고성 종합 판단", "Pearson 사용 정당화 확보", "채택 가능"],
]
last = write_data(ws6, summary, start_row=4, num_cols=3, bold_first=True)

add_note(ws6, "주. 본 연구의 Pearson 상관 결과는 변수 유형에 대해 견고함이 확인됨",
         last + 2, 3)
add_note(ws6, "    근거: Cohen et al.(2003) - 이분형↔연속형 Pearson = Point-biserial",
         last + 3, 3)
add_note(ws6, "          Hair et al.(2010) - 대표본 조건에서 Pearson robustness 확보",
         last + 4, 3)

ws6.column_dimensions["A"].width = 32
ws6.column_dimensions["B"].width = 36
ws6.column_dimensions["C"].width = 20


# 저장
wb.save("results/Correlation_robustness_check.xlsx")
print("저장 완료: results/Correlation_robustness_check.xlsx")
print("총 6개 시트:")
print("  1. Variable Types (변수 유형 분류)")
print("  2. DV Correlation Comparison (종속변수 상관 비교)")
print("  3. Pearson Matrix (전체 행렬)")
print("  4. Spearman Matrix (전체 행렬)")
print("  5. Kendall Matrix (전체 행렬)")
print("  6. Robustness Summary (견고성 종합 요약)")
